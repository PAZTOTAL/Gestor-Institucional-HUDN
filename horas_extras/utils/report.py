import io
import calendar
from datetime import date, timedelta
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .holidays import festivos_colombia


@dataclass
class EmpleadoInfo:
    """Datos del empleado necesarios para generar la planilla."""
    id: int
    nombre: str
    documento: str
    cargo: str
    area_nombre: str
    tipo: str           # 'permanente' | 'temporal'

    def get_tipo_display(self):
        return 'Planta Permanente' if self.tipo == 'permanente' else 'Planta Temporal'

    @classmethod
    def desde_dict(cls, d: dict) -> 'EmpleadoInfo':
        return cls(
            id=d['id'],
            nombre=d['nombre'],
            documento=d['documento'],
            cargo=d.get('cargo', ''),
            area_nombre=d.get('area_nombre', ''),
            tipo=d.get('tipo', 'permanente'),
        )

DIAS_ES = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
MESES_ES = [
    '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
]

# Horas (diurnas, nocturnas) por turno
# Nocturno: 18:00–06:00  |  Diurno: 06:00–18:00
TURNOS_HORAS = {
    'manana':        (6,  0),   # 07:00–13:00 → 6h diurnas
    'tarde':         (6,  0),   # 13:00–19:00 → 6h diurnas
    'noche':         (1, 11),   # 19:00–07:00 → 11h nocturnas + 1h diurna (06–07)
    'manana_noche':  (7, 11),   # 07:00–13:00 + 19:00–07:00 → 7h diurnas + 11h nocturnas
    'manana_tarde':  (12, 0),   # 07:00–19:00 → 12h diurnas
    'veinticuatro':  (12, 12),  # 07:00–07:00 → 12h diurnas + 12h nocturnas
    'libre':         (0,  0),   # sin cálculo
    'vacaciones':    (0,  0),   # sin horas laboradas
    'licencia':      (0,  0),   # sin horas laboradas
}

TURNOS_LABEL = {
    'manana':        'Mañana (07:00–13:00)',
    'tarde':         'Tarde (13:00–19:00)',
    'noche':         'Noche (19:00–07:00)',
    'manana_noche':  'Mañana-Noche (07:00–13:00 / 19:00–07:00)',
    'manana_tarde':  'Mañana-Tarde (07:00–19:00)',
    'veinticuatro':  '24 Horas (07:00–07:00)',
    'por_horas':     'Por horas',
    'libre':         'Libre',
    'vacaciones':    'V - Vacaciones',
    'licencia':      'LI - Licencia',
}

# Colores
C_HEADER  = '1F4E79'
C_TOTAL   = '1F4E79'
C_FESTIVO = 'FFE0E0'
C_DOMINGO = 'FFF0F0'
C_RESUMEN = 'D6E4F0'


def _border():
    t = Side(style='thin', color='BFBFBF')
    return Border(left=t, right=t, top=t, bottom=t)


def _fill(color):
    return PatternFill('solid', fgColor=color)


def _font(bold=False, color='000000', size=10, italic=False):
    return Font(name='Calibri', bold=bold, color=color, size=size, italic=italic)


def _es_siguiente_especial(fecha_obj, festivos):
    """True si el día siguiente al turno es domingo o festivo (cruza medianoche)."""
    siguiente = fecha_obj + timedelta(days=1)
    sig_fstr  = str(siguiente)
    if siguiente.year != fecha_obj.year:
        # Cruce de año: cargar festivos del año siguiente dinámicamente
        festivos_sig = festivos_colombia(siguiente.year)
        return siguiente.weekday() == 6 or sig_fstr in festivos_sig
    return siguiente.weekday() == 6 or sig_fstr in festivos


def calcular_horas(turno, es_festivo, horas_diurnas=0, horas_nocturnas=0,
                   siguiente_es_festivo=False, cross_month=False):
    """
    Retorna (hod, hon, hdf, hnf).

    siguiente_es_festivo: el día siguiente es domingo o festivo.
    cross_month: True cuando el turno es el último día del mes y el siguiente
                 día pertenece a otro mes. En ese caso solo se cuentan las horas
                 que ocurren ANTES de la medianoche (dentro del mes actual).
                 Las horas del día siguiente se calculan con calcular_spillover().
    """
    if turno in ('libre', 'vacaciones', 'licencia'):
        return 0, 0, 0, 0

    if turno == 'por_horas':
        hd, hn = int(horas_diurnas or 0), int(horas_nocturnas or 0)
        return (0, 0, hd, hn) if es_festivo else (hd, hn, 0, 0)

    # Turnos que cruzan medianoche
    if turno == 'noche':
        # 19:00–00:00 → mes actual  |  00:00–07:00 → día siguiente
        if cross_month and siguiente_es_festivo:
            # Solo horas hasta medianoche (19:00–00:00 = 5 h)
            return (0, 0, 0, 5) if es_festivo else (0, 5, 0, 0)
        if es_festivo and siguiente_es_festivo:
            return 0, 0, 1, 11        # todo festivo  →  1 hdf + 11 hnf
        if es_festivo and not siguiente_es_festivo:
            return 1, 6, 0, 5         # fest→normal   →  5 hnf + 6 hon + 1 hod
        if not es_festivo and siguiente_es_festivo:
            return 0, 5, 1, 6         # normal→fest   →  5 hon + 6 hnf + 1 hdf
        return 1, 11, 0, 0            # normal→normal →  1 hod + 11 hon

    if turno == 'manana_noche':
        # Mañana (07:00–13:00) + Noche (19:00–07:00)
        if cross_month and siguiente_es_festivo:
            # Mañana completa + noche solo hasta medianoche
            return (0, 0, 6, 5) if es_festivo else (6, 5, 0, 0)
        if es_festivo and siguiente_es_festivo:
            return 0, 0, 7, 11
        if es_festivo and not siguiente_es_festivo:
            return 1, 6, 6, 5
        if not es_festivo and siguiente_es_festivo:
            return 7, 5, 1, 6
        return 7, 11, 0, 0

    if turno == 'veinticuatro':
        # 07:00–00:00 → mes actual  |  00:00–07:00 → día siguiente
        if cross_month and siguiente_es_festivo:
            # Solo horas hasta medianoche (07:00–00:00 = 17 h → 11 hod + 6 hon)
            return (0, 0, 11, 6) if es_festivo else (11, 6, 0, 0)
        if es_festivo and siguiente_es_festivo:
            return 0, 0, 12, 12
        if es_festivo and not siguiente_es_festivo:
            return 1, 6, 11, 6
        if not es_festivo and siguiente_es_festivo:
            return 11, 6, 1, 6
        return 12, 12, 0, 0

    hd, hn = TURNOS_HORAS.get(turno, (0, 0))
    return (0, 0, hd, hn) if es_festivo else (hd, hn, 0, 0)


def calcular_spillover(turno, es_festivo, siguiente_es_festivo):
    """
    Horas del turno que caen después de medianoche en el día SIGUIENTE
    (solo relevante cuando el turno es el último día del mes y el primer
    día del mes siguiente es festivo/domingo).
    Retorna (hod, hon, hdf, hnf) que deben sumarse al mes siguiente.
    """
    if turno not in ('noche', 'manana_noche', 'veinticuatro'):
        return 0, 0, 0, 0
    if not siguiente_es_festivo:
        return 0, 0, 0, 0
    # 00:00–06:00 nocturno festivo + 06:00–07:00 diurno festivo
    return 0, 0, 1, 6


def _escribir_hoja_empleado(ws, trabajador, year, month, turnos_dict, festivos):
    """Rellena una hoja de cálculo con la planilla individual del empleado."""
    num_dias = calendar.monthrange(year, month)[1]

    brd = _border()
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center')

    # Título
    ws.merge_cells('A1:I1')
    ws['A1'] = 'PLANILLA DE TURNOS Y RECARGOS'
    ws['A1'].font = _font(bold=True, color=C_HEADER, size=14)
    ws['A1'].alignment = center

    # Info trabajador
    ws.merge_cells('A2:I2')
    ws['A2'] = (f"Trabajador: {trabajador.nombre}  |  "
                f"Cargo: {trabajador.cargo or '—'}  |  "
                f"Documento: {trabajador.documento}  |  "
                f"Tipo: {trabajador.get_tipo_display()}  |  "
                f"Área: {trabajador.area_nombre}")
    ws['A2'].font = _font(bold=True, size=11)
    ws['A2'].alignment = left

    ws.merge_cells('A3:I3')
    ws['A3'] = f"Período: {MESES_ES[month]} {year}"
    ws['A3'].font = _font(size=11)
    ws['A3'].alignment = left

    ws.append([])  # fila 4

    # Encabezados
    headers = [
        'Día', 'Fecha', 'Día Semana', 'Tipo Día',
        'Turno',
        'H. Ord. Diurnas', 'H. Ord. Nocturnas',
        'H. Diurnas Festivas', 'H. Nocturnas Festivas'
    ]
    ws.append(headers)
    hrow = ws.max_row
    for col in range(1, 10):
        c = ws.cell(row=hrow, column=col)
        c.font   = _font(bold=True, color='FFFFFF', size=11)
        c.fill   = _fill(C_HEADER)
        c.alignment = center
        c.border = brd
    ws.row_dimensions[hrow].height = 35

    totales = [0, 0, 0, 0]

    # ── Spillover del mes anterior ────────────────────────────────────────────
    # Si el primer día de este mes es festivo/domingo, verificar si el último
    # día del mes anterior tiene un turno que cruza medianoche hacia este mes.
    primer_dia   = date(year, month, 1)
    dia_anterior = primer_dia - timedelta(days=1)
    fstr_ant     = str(dia_anterior)
    t_ant = turnos_dict.get(fstr_ant)
    if t_ant is not None:
        t_ant_code = t_ant[0] if isinstance(t_ant, tuple) else t_ant
        primer_esp = (primer_dia.weekday() == 6 or str(primer_dia) in festivos)
        if primer_esp and t_ant_code in ('noche', 'manana_noche', 'veinticuatro'):
            ant_fest  = (dia_anterior.weekday() == 6 or fstr_ant in festivos)
            sp = calcular_spillover(t_ant_code, ant_fest, primer_esp)
            if any(sp):
                totales[0] += sp[0]; totales[1] += sp[1]
                totales[2] += sp[2]; totales[3] += sp[3]
                # Fila informativa de spillover
                ws.merge_cells(f'A{ws.max_row + 1}:I{ws.max_row + 1}')
                nota_row = ws.max_row
                ws.cell(row=nota_row, column=1,
                        value=(f'⚠ Horas adicionales acreditadas al {MESES_ES[month]}: '
                               f'{sp[2]} HDF + {sp[3]} HNF provenientes del turno '
                               f'{TURNOS_LABEL.get(t_ant_code, t_ant_code)} '
                               f'del {dia_anterior.strftime("%d/%m/%Y")} '
                               f'(horas después de medianoche en día festivo/domingo).'))
                ws.cell(row=nota_row, column=1).font  = _font(bold=True, color='7B341E', size=9)
                ws.cell(row=nota_row, column=1).fill  = _fill('FFF3CD')
                ws.append([])

    for dia in range(1, num_dias + 1):
        fecha    = date(year, month, dia)
        fstr     = str(fecha)
        es_dom   = fecha.weekday() == 6
        es_fest  = fstr in festivos
        especial = es_dom or es_fest
        sig_esp  = _es_siguiente_especial(fecha, festivos)

        # Detectar último día del mes con turno que cruza hacia el mes siguiente
        es_ultimo_dia = (dia == num_dias)
        siguiente_fecha = fecha + timedelta(days=1)
        cross_month = es_ultimo_dia and sig_esp and (siguiente_fecha.month != month)

        tipo_dia = (f"Festivo ({festivos[fstr]})" if es_fest
                    else "Domingo" if es_dom else "Ordinario")

        t_info = turnos_dict.get(fstr)
        if t_info is None:
            t_code = None
            t_hd = t_hn = 0
        elif isinstance(t_info, tuple):
            t_code, t_hd, t_hn = t_info
        else:
            t_code, t_hd, t_hn = t_info, 0, 0

        t_label = TURNOS_LABEL.get(t_code, '—') if t_code else '—'

        if t_code:
            hod, hon, hdf, hnf = calcular_horas(
                t_code, especial, t_hd, t_hn, sig_esp, cross_month=cross_month)
            totales[0] += hod; totales[1] += hon
            totales[2] += hdf; totales[3] += hnf
        else:
            hod = hon = hdf = hnf = ''

        ws.append([dia, fecha.strftime('%d/%m/%Y'), DIAS_ES[fecha.weekday()],
                   tipo_dia, t_label, hod, hon, hdf, hnf])
        drow = ws.max_row

        for col in range(1, 10):
            c = ws.cell(row=drow, column=col)
            c.font   = _font()
            c.border = brd
            c.alignment = center

            if es_fest:   c.fill = _fill('FFE0E0')
            elif es_dom:  c.fill = _fill('FFF0F0')

        if especial:
            for col in [1, 3, 4]:
                ws.cell(row=drow, column=col).font = _font(bold=True, color='C53030')

        # Nota de cross-month al pie del último día
        if cross_month and t_code in ('noche', 'manana_noche', 'veinticuatro'):
            sp = calcular_spillover(t_code, especial, sig_esp)
            if any(sp):
                ws.append(['', '', '', '', '',
                           f'↳ {sp[2]} HDF + {sp[3]} HNF se acreditan a '
                           f'{MESES_ES[siguiente_fecha.month]} '
                           f'(horas después de medianoche en día festivo/domingo)'])
                nrow = ws.max_row
                ws.merge_cells(f'F{nrow}:I{nrow}')
                ws.cell(row=nrow, column=6).font      = _font(italic=True, color='7B341E', size=9)
                ws.cell(row=nrow, column=6).fill      = _fill('FFF3CD')
                ws.cell(row=nrow, column=6).alignment = Alignment(horizontal='left')

    # Fila totales
    ws.append([])
    ws.append(['', '', '', 'TOTALES', '', totales[0], totales[1], totales[2], totales[3]])
    trow = ws.max_row
    for col in range(1, 10):
        c = ws.cell(row=trow, column=col)
        c.font = _font(bold=True, color='FFFFFF', size=11)
        c.fill = _fill(C_TOTAL)
        c.alignment = center
        c.border = brd

    # Leyenda
    ws.append([])
    for txt in [
        'LEYENDA:',
        'HOD = Horas Ordinarias Diurnas (06:00–21:00)',
        'HON = Horas Ordinarias Nocturnas (21:00–06:00) — Recargo 35%',
        'HDF = Horas Diurnas Festivas (domingos/festivos diurnas) — Recargo 75%',
        'HNF = Horas Nocturnas Festivas (domingos/festivos nocturnas) — Recargo 110%',
    ]:
        ws.append([txt])
        ws.cell(row=ws.max_row, column=1).font = _font(italic=True, size=9, color='595959')

    # Anchos
    for i, w in enumerate([6, 14, 14, 30, 26, 18, 20, 20, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return totales


def generar_planilla(trabajador, year, month, turnos_dict):
    """Genera planilla individual. Retorna bytes Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{MESES_ES[month]} {year}"[:31]

    festivos = festivos_colombia(year)
    _escribir_hoja_empleado(ws, trabajador, year, month, turnos_dict, festivos)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generar_planilla_area(area, year, month, empleados_turnos, coordinador_nombre=''):
    """
    Genera planilla del área completa.
    empleados_turnos: list of (trabajador, turnos_dict)
    Retorna bytes Excel.
    """
    wb = Workbook()
    festivos = festivos_colombia(year)

    # ── Sheet 1: Resumen ──────────────────────────────────────
    ws_res = wb.active
    ws_res.title = 'Resumen'

    brd    = _border()
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center')

    ws_res.merge_cells('A1:H1')
    ws_res['A1'] = f'PLANILLA DEL ÁREA: {area.nombre.upper()} — {MESES_ES[month]} {year}'
    ws_res['A1'].font      = _font(bold=True, color=C_HEADER, size=14)
    ws_res['A1'].alignment = center

    if coordinador_nombre:
        ws_res.merge_cells('A2:H2')
        ws_res['A2'] = f'Coordinador del servicio: {coordinador_nombre}'
        ws_res['A2'].font      = _font(bold=False, size=11)
        ws_res['A2'].alignment = left
        ws_res.append([])
    else:
        ws_res.append([])

    headers = ['Empleado', 'Documento', 'Cargo', 'Tipo',
               'H. Ord. Diurnas', 'H. Ord. Nocturnas',
               'H. Diurnas Festivas', 'H. Nocturnas Festivas']
    ws_res.append(headers)
    hrow = ws_res.max_row
    for col in range(1, 9):
        c = ws_res.cell(row=hrow, column=col)
        c.font      = _font(bold=True, color='FFFFFF', size=11)
        c.fill      = _fill(C_HEADER)
        c.alignment = center
        c.border    = brd
    ws_res.row_dimensions[hrow].height = 35

    tot_global = [0, 0, 0, 0]
    tipo_actual = None

    for trabajador, turnos_dict in empleados_turnos:
        # Separador de tipo
        if trabajador.tipo != tipo_actual:
            tipo_actual = trabajador.tipo
            ws_res.append([trabajador.get_tipo_display()])
            srow = ws_res.max_row
            ws_res.merge_cells(f'A{srow}:H{srow}')
            ws_res.cell(row=srow, column=1).font      = _font(bold=True, color='FFFFFF', size=10)
            ws_res.cell(row=srow, column=1).fill      = _fill('2D6FAD')
            ws_res.cell(row=srow, column=1).alignment = left
            for col in range(1, 9):
                ws_res.cell(row=srow, column=col).border = brd

        # Calcular totales del empleado
        hod = hon = hdf = hnf = 0
        for fstr, t_info in turnos_dict.items():
            if isinstance(t_info, tuple):
                t_code, t_hd, t_hn = t_info
            else:
                t_code, t_hd, t_hn = t_info, 0, 0
            fecha    = date.fromisoformat(fstr)
            especial = fecha.weekday() == 6 or fstr in festivos
            sig_esp  = _es_siguiente_especial(fecha, festivos)
            h = calcular_horas(t_code, especial, t_hd, t_hn, sig_esp)
            hod += h[0]; hon += h[1]; hdf += h[2]; hnf += h[3]

        tot_global[0] += hod; tot_global[1] += hon
        tot_global[2] += hdf; tot_global[3] += hnf

        ws_res.append([trabajador.nombre, trabajador.documento,
                       trabajador.cargo or '—', trabajador.get_tipo_display(),
                       hod, hon, hdf, hnf])
        drow = ws_res.max_row
        for col in range(1, 9):
            c = ws_res.cell(row=drow, column=col)
            c.font      = _font()
            c.border    = brd
            c.alignment = center if col >= 5 else left

    # Total global
    ws_res.append([])
    ws_res.append(['', '', '', 'TOTALES ÁREA',
                   tot_global[0], tot_global[1], tot_global[2], tot_global[3]])
    trow = ws_res.max_row
    for col in range(1, 9):
        c = ws_res.cell(row=trow, column=col)
        c.font      = _font(bold=True, color='FFFFFF', size=11)
        c.fill      = _fill(C_TOTAL)
        c.alignment = center
        c.border    = brd

    # Anchos resumen
    for i, w in enumerate([30, 16, 22, 18, 18, 20, 20, 22], 1):
        ws_res.column_dimensions[get_column_letter(i)].width = w

    # ── Sheets individuales ───────────────────────────────────
    for trabajador, turnos_dict in empleados_turnos:
        nombre_sheet = trabajador.nombre[:28].strip()
        ws_emp = wb.create_sheet(title=nombre_sheet)
        _escribir_hoja_empleado(ws_emp, trabajador, year, month, turnos_dict, festivos)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
