"""
Script directo con pyodbc para:
1. Crear la tabla defenjur_app_despachojudicial en FENJUR_BD_ORG
2. Cargar los 55 registros del Excel defenjur_app_despachojudicial.xlsx

Ejecutar desde: Gestor Institucional HUDN/defenjur_py/
"""
import pyodbc
import openpyxl
import os

CONN_STR = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=localhost\\SQLEXPRESS;"
    "DATABASE=FENJUR_BD_ORG;"
    "Trusted_Connection=yes;"
    "TrustServerCertificate=yes;"
)

EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "despachoJudicial.xlsx"
)

CREATE_TABLE_SQL = """
IF NOT EXISTS (
    SELECT * FROM INFORMATION_SCHEMA.TABLES 
    WHERE TABLE_NAME = 'defenjur_app_despachojudicial'
)
BEGIN
    CREATE TABLE defenjur_app_despachojudicial (
        id       INT IDENTITY(1,1) PRIMARY KEY,
        ciudad   NVARCHAR(100)  NOT NULL,
        nombre   NVARCHAR(255)  NOT NULL,
        correo   NVARCHAR(255)  NULL
    );
    PRINT 'Tabla defenjur_app_despachojudicial creada.';
END
ELSE
BEGIN
    PRINT 'La tabla defenjur_app_despachojudicial ya existe.';
END
"""

INSERT_SQL = """
INSERT INTO defenjur_app_despachojudicial (ciudad, nombre, correo) VALUES (?, ?, ?)
"""

def main():
    print(f"Excel path: {EXCEL_PATH}")
    if not os.path.exists(EXCEL_PATH):
        print("ERROR: No se encontró el archivo Excel.")
        return

    # Leer Excel
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    rows = []
    for row_num in range(2, ws.max_row + 1):
        no     = ws.cell(row_num, 1).value
        ciudad = ws.cell(row_num, 2).value
        nombre = ws.cell(row_num, 3).value
        correo = ws.cell(row_num, 4).value
        if nombre:  # Omitir filas vacías
            rows.append((
                str(ciudad).strip() if ciudad else '',
                str(nombre).strip(),
                str(correo).strip() if correo else None
            ))

    print(f"Registros leídos del Excel: {len(rows)}")

    # Conectar a SQL Server
    try:
        conn = pyodbc.connect(CONN_STR)
        cursor = conn.cursor()
        print("Conexión exitosa a FENJUR_BD_ORG.")

        # Crear tabla si no existe
        cursor.execute(CREATE_TABLE_SQL)
        conn.commit()

        # Verificar si ya tiene datos
        cursor.execute("SELECT COUNT(*) FROM defenjur_app_despachojudicial")
        count = cursor.fetchone()[0]

        if count > 0:
            print(f"La tabla ya tiene {count} registros. Limpiando para recargar...")
            cursor.execute("DELETE FROM defenjur_app_despachojudicial")
            conn.commit()

        # Insertar registros
        cursor.executemany(INSERT_SQL, rows)
        conn.commit()

        # Verificar
        cursor.execute("SELECT COUNT(*) FROM defenjur_app_despachojudicial")
        total = cursor.fetchone()[0]
        print(f"\n✅ Carga completada. Total de registros en la tabla: {total}")

        # Mostrar primeros 5
        cursor.execute("SELECT TOP 5 id, ciudad, nombre, correo FROM defenjur_app_despachojudicial ORDER BY ciudad, nombre")
        print("\nPrimeros 5 registros cargados:")
        for r in cursor.fetchall():
            print(f"  [{r[0]}] {r[1]} | {r[2]} | {r[3]}")

        cursor.close()
        conn.close()

    except Exception as e:
        print(f"\n❌ Error: {e}")

if __name__ == "__main__":
    main()
