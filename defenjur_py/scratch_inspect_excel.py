import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\SISTEMAS\Documents\Gestor Institucional HUDN\despachoJudicial.xlsx")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Hoja: {sheet_name} ===")
    print(f"Dimensiones: {ws.dimensions} | Filas: {ws.max_row} | Columnas: {ws.max_column}")
    print("\nEncabezados (fila 1):")
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    for i, h in enumerate(headers, 1):
        print(f"  Col {i}: {h}")
    print("\nPrimeras 5 filas de datos:")
    for row in range(2, min(7, ws.max_row + 1)):
        row_data = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        print(f"  Fila {row}: {row_data}")
