from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.urls import reverse_lazy
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.utils import timezone
import pandas as pd
from datetime import datetime
from .models import (
    MezclaOrden, MezclaPreparacion, MezclaControlCalidad, MezclaDistribucion,
    ReempaqueMedicamento, ReempaqueOrden, ReempaqueControl, ReempaqueMuestreo,
    Alerta, ConvencionFormaFarmaceutica, MedicamentoEsteril,
    UnidosisPeriodo, UnidosisOrden, MedicamentoOncologico,
    OncologicoMatriz, OncologicoMatrizItem,
    OncologicoOrdenProduccion, OncologicoOrdenItem,
    OncologicoAlistamiento, OncologicoAlistamientoItem,
    NeonatosMatriz, NeonatosMatrizItem, NeonatosMedicamento, NeonatosOrdenProduccion,
    NeonatosOrdenItem, NeonatosAlistamiento, NeonatosAlistamientoItem,
    UnidosisProduccionOrden, NptMatriz, NptMatrizItem, NptOrdenProduccion, NptOrdenItem, NptAlistamiento, NptAlistamientoItem,
    Funcionario
)
from consultas_externas.models import (
    Genpacien, Adningreso, Hcnfolio, Hcnmedpac, Innproduc, Hpndefcam, 
    Genareser, Hpnestanc, Gendetcon, Innfarcol, Gentercer, Comtercero
)
from HospitalManagement.api_views import calculate_age
from django.db.models import Q
from core.mixins import SearchFilterMixin
from django import forms
import json
from django.views.decorators.http import require_http_methods
import io
from xhtml2pdf import pisa
from django.template.loader import get_template
import openpyxl
from django.views.decorators.http import require_GET

class CentralDeMezclasMainView(LoginRequiredMixin, TemplateView):
    template_name = 'CentralDeMezclas/main_selector.html'

class EsterilesDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'CentralDeMezclas/esteriles/dashboard.html'

# --- VISTAS BD MEDICAMENTOS ESTÉRILES ---

class MedicamentoEsterilListView(LoginRequiredMixin, SearchFilterMixin, ListView):
    model = MedicamentoEsteril
    template_name = 'CentralDeMezclas/esteriles/medicamento_list.html'
    context_object_name = 'medicamentos'

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.GET.get('ace') == 'true':
            queryset = queryset.filter(es_control_especial=True)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_ace'] = self.request.GET.get('ace') == 'true'
        return context

class MedicamentoEsterilCreateView(LoginRequiredMixin, CreateView):
    model = MedicamentoEsteril
    template_name = 'CentralDeMezclas/esteriles/medicamento_form.html'
    fields = '__all__'
    success_url = reverse_lazy('mezclas:esteriles_med_list')

class MedicamentoEsterilUpdateView(LoginRequiredMixin, UpdateView):
    model = MedicamentoEsteril
    template_name = 'CentralDeMezclas/esteriles/medicamento_form.html'
    fields = '__all__'
    success_url = reverse_lazy('mezclas:esteriles_med_list')

class MedicamentoEsterilDeleteView(LoginRequiredMixin, DeleteView):
    model = MedicamentoEsteril
    template_name = 'CentralDeMezclas/confirm_delete.html'
    success_url = reverse_lazy('mezclas:esteriles_med_list')

# --- MATRIZ UNIDOSIS ADULTOS ---

class UnidosisPeriodoListView(LoginRequiredMixin, SearchFilterMixin, ListView):
    model = UnidosisPeriodo
    template_name = 'CentralDeMezclas/esteriles/unidosis_list.html'
    context_object_name = 'periodos'

class UnidosisPeriodoCreateView(LoginRequiredMixin, CreateView):
    model = UnidosisPeriodo
    template_name = 'CentralDeMezclas/esteriles/unidosis_periodo_form.html'
    fields = '__all__'
    
    def get_success_url(self):
        return reverse_lazy('mezclas:unidosis_matriz', kwargs={'pk': self.object.pk})

class UnidosisMatrizView(LoginRequiredMixin, DetailView):
    model = UnidosisPeriodo
    template_name = 'CentralDeMezclas/esteriles/unidosis_matriz.html'
    context_object_name = 'periodo'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['ordenes'] = self.object.ordenes.all()
        return context

class UnidosisOrdenCreateView(LoginRequiredMixin, CreateView):
    model = UnidosisOrden
    template_name = 'CentralDeMezclas/esteriles/unidosis_orden_form.html'
    fields = '__all__'

    def get_initial(self):
        initial = super().get_initial()
        periodo_id = self.kwargs.get('periodo_id')
        if periodo_id:
            initial['periodo'] = periodo_id
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['periodo_id'] = self.kwargs.get('periodo_id')
        return context

    def get_success_url(self):
        return reverse_lazy('mezclas:unidosis_matriz', kwargs={'pk': self.object.periodo.pk})

class ReempaquePanelView(LoginRequiredMixin, TemplateView):
    template_name = 'CentralDeMezclas/reempaque/panel.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['stats'] = {
            'medicamentos': ReempaqueMedicamento.objects.count(),
            'matrices': ReempaqueOrden.objects.count(),
            'ordenes': ReempaqueOrden.objects.count(),
            'alistamientos': 0, # Próximamente
        }
        return context

# --- VISTAS DE MEZCLAS (MAGISTRALES) ---
class MezclaOrdenCreateView(LoginRequiredMixin, CreateView):
    model = MezclaOrden
    template_name = 'CentralDeMezclas/orden_form.html'
    fields = ['paciente_oid', 'medico_oid', 'tipo_mezcla', 'descripcion_medicamento', 'prioridad', 'observaciones_clinicas']
    success_url = reverse_lazy('mezclas:reempaque_panel')

    def form_valid(self, form):
        return super().form_valid(form)

class MezclaOrdenDetailView(LoginRequiredMixin, DetailView):
    model = MezclaOrden
    template_name = 'CentralDeMezclas/orden_detail.html'
    context_object_name = 'orden'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        try:
            context['paciente'] = Genpacien.objects.using('readonly').get(oid=self.object.paciente_oid)
        except Genpacien.DoesNotExist:
            context['paciente'] = None
        return context

# --- VISTAS DE REEMPAQUE Y REENVASE (RV-RS) ---

class ReempaqueMedicamentoListView(LoginRequiredMixin, SearchFilterMixin, ListView):
    model = ReempaqueMedicamento
    template_name = 'CentralDeMezclas/reempaque/medicamento_list.html'
    context_object_name = 'medicamentos'
class ReempaqueMedicamentoCreateView(LoginRequiredMixin, CreateView):
    model = ReempaqueMedicamento
    template_name = 'CentralDeMezclas/reempaque/medicamento_form.html'
    fields = '__all__'
    success_url = reverse_lazy('mezclas:reempaque_med_list')

class ReempaqueMedicamentoUpdateView(LoginRequiredMixin, UpdateView):
    model = ReempaqueMedicamento
    template_name = 'CentralDeMezclas/reempaque/medicamento_form.html'
    fields = '__all__'
    success_url = reverse_lazy('mezclas:reempaque_med_list')

class ReempaqueMedicamentoDeleteView(LoginRequiredMixin, DeleteView):
    model = ReempaqueMedicamento
    template_name = 'CentralDeMezclas/confirm_delete.html'
    success_url = reverse_lazy('mezclas:reempaque_med_list')

    success_url = reverse_lazy('mezclas:reempaque_matriz')

class ReempaqueOrdenCreateView(LoginRequiredMixin, CreateView):
    model = ReempaqueOrden
    template_name = 'CentralDeMezclas/reempaque/orden_form.html'
    fields = [
        'tipo', 'medicamento', 'lote_fabricante', 'laboratorio', 'registro_invima', 
        'fecha_vencimiento', 'cantidad_a_reempacar', 'semaforizacion', 
        'responsable_reempaque', 'farmaceutico_validador'
    ]
    success_url = reverse_lazy('mezclas:reempaque_matriz')

    def get_initial(self):
        initial = super().get_initial()
        # Valores por defecto para agilizar el registro
        initial['farmaceutico_validador'] = self.request.user.id
        return initial

    def form_valid(self, form):
        # Lógica REEHUDN1-YY-MM-####
        now = timezone.now()
        prefix = f"REEHUDN1-{now.strftime('%y-%m')}"
        count = ReempaqueOrden.objects.filter(lote_interno__startswith=prefix).count() + 1
        form.instance.lote_interno = f"{prefix}-{count:04d}"
        
        # El vencimiento de reempaque suele ser 1 año o el del fabricante (el menor)
        form.instance.fecha_vencimiento_reempaque = form.instance.fecha_vencimiento
        
        return super().form_valid(form)

class ReempaqueOrdenUpdateView(LoginRequiredMixin, UpdateView):
    model = ReempaqueOrden
    template_name = 'CentralDeMezclas/reempaque/orden_form.html'
    fields = [
        'tipo', 'medicamento', 'lote_fabricante', 'laboratorio', 'registro_invima', 
        'fecha_vencimiento', 'cantidad_a_reempacar', 'semaforizacion', 
        'responsable_reempaque', 'farmaceutico_validador'
    ]
    success_url = reverse_lazy('mezclas:reempaque_matriz')

class ReempaqueOrdenDeleteView(LoginRequiredMixin, DeleteView):
    model = ReempaqueOrden
    template_name = 'CentralDeMezclas/confirm_delete.html'
    success_url = reverse_lazy('mezclas:reempaque_matriz')

class ReempaqueEtiquetaView(LoginRequiredMixin, DetailView):
    """Generación de etiquetas (FRFAR-072)"""
    model = ReempaqueOrden
    template_name = 'CentralDeMezclas/reempaque/etiqueta_print.html'
    context_object_name = 'orden'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tipo'] = self.request.GET.get('tipo', 'BN')
        return context

class ReempaqueControlCreateView(LoginRequiredMixin, CreateView):
    """Control de Calidad (FRFAR-188)"""
    model = ReempaqueControl
    template_name = 'CentralDeMezclas/reempaque/control_form.html'
    fields = [
        'aspecto_fisico_ok', 'hermeticidad_ok', 'etiquetado_completo_ok', 
        'limpieza_area_ok', 'unidades_defectuosas', 'observaciones_calidad'
    ]

    def form_valid(self, form):
        orden = get_object_or_404(ReempaqueOrden, id=self.kwargs['orden_id'])
        form.instance.orden = orden
        
        # Generar Muestreo Automático (MIL-STD-105E)
        plan = ReempaqueMuestreo.obtener_plan(orden.cantidad_a_reempacar)
        ReempaqueMuestreo.objects.create(
            orden=orden,
            tamano_lote=orden.cantidad_a_reempacar,
            letra_codigo=plan[0],
            tamano_muestra=plan[1],
            ac=plan[2],
            re=plan[3]
        )

        # Si todo está OK y no hay unidades defectuosas críticas (menor al Re del plan)
        if all([form.cleaned_data['aspecto_fisico_ok'], 
                form.cleaned_data['hermeticidad_ok'], 
                form.cleaned_data['etiquetado_completo_ok']]) and form.cleaned_data['unidades_defectuosas'] < plan[3]:
            form.instance.liberado = True
            # Asignar responsable de liberación si existe el funcionario asociado al usuario
            try:
                funcionario = Funcionario.objects.get(cedula=self.request.user.username)
                form.instance.responsable_liberacion = funcionario
            except Funcionario.DoesNotExist:
                pass
            form.instance.fecha_liberacion = timezone.now()
            orden.estado = 'LIBERADO'
            orden.cantidad_final_aprobada = orden.cantidad_a_reempacar - form.cleaned_data['unidades_defectuosas']
        else:
            orden.estado = 'RECHAZADO'
        
        orden.save()
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('mezclas:reempaque_matriz')

class ReempaqueMatrizView(LoginRequiredMixin, SearchFilterMixin, ListView):
    """La Matriz Operativa (FRFAR-090)"""
    model = ReempaqueOrden
    template_name = 'CentralDeMezclas/reempaque/matriz.html'
    context_object_name = 'ordenes'
    
    def get_queryset(self):
        return ReempaqueOrden.objects.all().order_by('-fecha_produccion')

class ReempaqueAlistamientoView(LoginRequiredMixin, TemplateView):
    """Alistamiento y Conciliación (FRFAR-162)"""
    template_name = 'CentralDeMezclas/reempaque/alistamiento.html'

class ReempaqueMuestreoView(LoginRequiredMixin, TemplateView):
    """Muestreo de Medicamentos (FRFAR-092)"""
    template_name = 'CentralDeMezclas/reempaque/muestreo.html'

class ReempaqueNormaView(LoginRequiredMixin, TemplateView):
    """Norma MIL-STD-105E (PTFAR-005-AN-02)"""
    template_name = 'CentralDeMezclas/reempaque/norma.html'

def buscar_pacientes(request):
    query = request.GET.get('q', '')
    if len(query) < 3:
        return JsonResponse({'results': []})
    
    pacientes = Genpacien.objects.using('readonly').filter(
        Q(pacnumdoc__icontains=query) | 
        Q(pacprinom__icontains=query) | 
        Q(pacpriape__icontains=query)
    )[:10]
    
    results = []
    for p in pacientes:
        nombre_completo = f"{p.pacprinom or ''} {p.pacsegnom or ''} {p.pacpriape or ''} {p.pacsegape or ''}".strip()
        results.append({
            'oid': p.oid,
            'doc': p.pacnumdoc,
            'nombre': nombre_completo,
            'text': f"{p.pacnumdoc} - {nombre_completo}"
        })
    
    return JsonResponse({'results': results})

def api_get_medicamento(request, pk):
    """Retorna los datos maestros de un medicamento (base madre)"""
    med = get_object_or_404(ReempaqueMedicamento, pk=pk)
    data = {
        'nombre': med.nombre,
        'concentracion': med.concentracion,
        'forma_farmaceutica': med.convencion.forma_farmaceutica if med.convencion else "N/A",
        'via': med.convencion.via if med.convencion else "N/A",
        'alerta_codigo': med.alerta.codigo if med.alerta else "N/A",
        'alerta_color': med.alerta.color if med.alerta else "#ccc",
        'condiciones_almacenamiento': med.condiciones_almacenamiento,
        'fotosensibilidad': med.fotosensibilidad,
    }
    return JsonResponse(data)

@login_required
@require_http_methods(["GET", "POST"])
def api_cargar_matriz_unidosis(request):
    """
    GET: Consulta pacientes activos en un rango de camas.
    POST: Guarda masivamente los resultados en el periodo actual.
    """
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            periodo_id = data.get('periodo_id')
            items = data.get('items', [])
            
            periodo = get_object_or_404(UnidosisPeriodo, pk=periodo_id)
            
            created_count = 0
            for item in items:
                # Generar LOTE INTERNO incremental para este periodo
                ultimo_lote = UnidosisOrden.objects.filter(periodo=periodo).count() + 1
                lote_str = f"UNI-{periodo.orden_produccion}-{ultimo_lote:03d}"
                
                UnidosisOrden.objects.create(
                    periodo=periodo,
                    lote_interno=lote_str,
                    paciente_nombre=item['paciente_nombre'],
                    paciente_identificacion=item['paciente_identificacion'],
                    cama=item['cama'],
                    servicio=item['servicio'],
                    medicamento_base_id=item['med_oid'],
                    dosis_estandar_pme=item['med_nombre_completo'],
                    dosis_frecuencia=item['dosis_frecuencia'],
                    sln_diluyente=item['diluyente'],
                    volumen_dilucion=item['vol_dilucion'],
                    vehiculo_volumen_final=item['vehiculo'],
                    via=item['via'],
                    frecuencia=item['dosis_frecuencia'].split()[-1] if ' ' in item['dosis_frecuencia'] else '24h',
                    cantidad_necesaria=item['cantidad'],
                    volumen_dosis=item['vol_dilucion'],
                    fecha_expiracion=timezone.now() + timezone.timedelta(days=1), # 24h default
                    temp_almacenamiento=item['temp'],
                    luz_almacenamiento=item['luz']
                )
                created_count += 1
                
            return JsonResponse({'status': 'ok', 'created': created_count})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    # Lógica GET (Búsqueda)
    cama_ini = request.GET.get('desde', '')
    cama_fin = request.GET.get('hasta', '')
    
    if not cama_ini or not cama_fin:
        return JsonResponse({'error': 'Debe proporcionar rango de camas'}, status=400)

    # 1. Buscar camas en el rango que tengan un ingreso activo
    camas_activas = Hpndefcam.objects.using('readonly').filter(
        hcacodigo__gte=cama_ini,
        hcacodigo__lte=cama_fin,
        adningreso__isnull=False
    ).select_related()

    results = []
    
    # Obtener códigos de medicamentos estériles permitidos (Maestro Técnico)
    maestro_esteriles = {}
    for m in MedicamentoEsteril.objects.all():
        if m.cod_med_1 and m.cod_med_1 != 'N/A':
            maestro_esteriles[m.cod_med_1] = m
        if m.cod_med_2 and m.cod_med_2 != '':
            maestro_esteriles[m.cod_med_2] = m
    
    codigos_maestro = maestro_esteriles.keys()

    for cama in camas_activas:
        # 2. Obtener el ingreso y el paciente
        try:
            ingreso = Adningreso.objects.using('readonly').get(oid=cama.adningreso)
            paciente = ingreso.genpacien
            
            # 3. Buscar el último folio abierto del paciente en este ingreso
            folio = Hcnfolio.objects.using('readonly').filter(
                adningreso=ingreso.oid
            ).order_by('-hcfecfol').first()

            if not folio:
                continue

            # 4. Buscar medicamentos prescritos en este folio que estén en el maestro
            prescripciones = Hcnmedpac.objects.using('readonly').filter(
                hcnfolio=folio.oid,
                hcsterdef=False, # No suspendido
                hcsfecsus__isnull=True
            )

            for p in prescripciones:
                # Obtener info del producto
                try:
                    producto = Innproduc.objects.using('readonly').get(oid=p.innproduc)
                    # Validar si el código del producto coincide con nuestro maestro
                    if producto.iprcodigo in codigos_maestro:
                        med_base = maestro_esteriles[producto.iprcodigo]
                        
                        nombre_paciente = f"{paciente.pacprinom or ''} {paciente.pacsegnom or ''} {paciente.pacpriape or ''} {paciente.pacsegape or ''}".strip()
                        
                        results.append({
                            'paciente_nombre': nombre_paciente,
                            'paciente_id': paciente.pacnumdoc,
                            'fecha_nacimiento': paciente.gpafecnac.strftime('%d/%m/%Y') if paciente.gpafecnac else "N/A",
                            'cama': cama.hcacodigo if cama else "S/C",
                            'servicio': cama.hcanombre,
                            'med_codigo': med_base.codigo,
                            'med_oid': med_base.pk,
                            'med_nombre_completo': f"{med_base.medicamento_1} {med_base.concentracion_1}",
                            'dosis_frecuencia': f"{p.hcsdosis or ''} cada {p.hcsfrecmed or ''}h",
                            'via': med_base.via,
                            'vehiculo': med_base.vehiculo_final,
                            'diluyente': med_base.diluyente,
                            'vol_dilucion': med_base.vol_reconstitucion_1,
                            'cantidad': str(p.hcscanti or '1'), # Cantidad de la fórmula
                            'temp': med_base.almacenamiento,
                            'luz': 'PROTEGER LUZ' if med_base.alto_riesgo else 'NORMAL',
                        })
                except Innproduc.DoesNotExist:
                    continue

        except Adningreso.DoesNotExist:
            continue

    return JsonResponse({'results': results, 'count': len(results)})

class ConsultaFormulasAreaView(LoginRequiredMixin, TemplateView):
    template_name = 'CentralDeMezclas/consultas/formulas_area.html'

@require_GET
@login_required
def api_buscar_tercero_hudn(request):
    """Busca un tercero (paciente o personal) en Gentercer o Comtercero por identificación."""
    identificacion = request.GET.get('identificacion', '')
    if not identificacion:
        return JsonResponse({'error': 'No se proporcionó identificación'}, status=400)
    
    # 1. Intentar en Gentercer (General HUDN)
    tercero = Gentercer.objects.using('readonly').filter(ternumdoc=identificacion).first()
    if tercero:
        nombre = tercero.ternomcom or f"{tercero.terprinom or ''} {tercero.tersegnom or ''} {tercero.terpriape or ''} {tercero.tersegape or ''}".strip()
        return JsonResponse({
            'success': True,
            'cedula': tercero.ternumdoc,
            'nombre': nombre,
            'origen': 'GENTERCER'
        })
    
    # 2. Fallback a Comtercero (Consultas Externas)
    tercero_com = Comtercero.objects.using('readonly').filter(teridenti=identificacion).first()
    if tercero_com:
        return JsonResponse({
            'success': True,
            'cedula': tercero_com.teridenti,
            'nombre': tercero_com.ternomcom,
            'origen': 'COMTERCERO'
        })

    return JsonResponse({'success': False, 'error': 'No encontrado en HUDN'}, status=404)

@require_GET
@login_required
def api_buscar_personal_hudn(request):
    """Busca personal en Gentercer para componentes Select2."""
    q = request.GET.get('q', '')
    results = []
    if q:
        # Búsqueda por documento o nombre completo
        qs = Gentercer.objects.using('readonly').filter(
            Q(ternumdoc__icontains=q) | 
            Q(ternomcom__icontains=q) |
            Q(terprinom__icontains=q) |
            Q(terpriape__icontains=q)
        ).order_by('ternomcom')[:25]
        
        for t in qs:
            nombre = t.ternomcom or f"{t.terprinom or ''} {t.terpriape or ''}".strip()
            results.append({
                'id': t.ternumdoc,
                'text': f"{t.ternumdoc} - {nombre}",
                'nombre': nombre
            })
            
    return JsonResponse({'results': results})

def api_formulas_por_area(request, area_id):
    """
    Retorna la lista de fórmulas de todos los pacientes activos en un área.
    Si request.GET.get('single') == 'true', area_id se trata como el ID del ingreso individual.
    """
    single_mode = request.GET.get('single')
    
    # 1. Obtener ingresos activos
    if single_mode == 'true':
        # Consulta por un único ingreso
        ingresos = Adningreso.objects.using('readonly').filter(
            oid=area_id,
            ainfecegre__isnull=True
        ).select_related('genpacien')
    elif single_mode == 'global':
        # Consulta GLOBAL: todos los ingresos activos en el sistema (Hospitalización y Urgencias)
        ingresos = Adningreso.objects.using('readonly').filter(
            ainfecegre__isnull=True,
            ainestado=1
        ).select_related('genpacien').order_by('hpndefcam')
    elif ',' in str(area_id):
        # Consulta por lista de IDs (ej: "123,456,789")
        ids = [i.strip() for i in str(area_id).split(',') if i.strip().isdigit()]
        ingresos = Adningreso.objects.using('readonly').filter(
            oid__in=ids,
            ainfecegre__isnull=True
        ).select_related('genpacien')
    else:
        # Consulta por área (genareser o adncenate)
        ingresos = Adningreso.objects.using('readonly').filter(
            Q(genareser=area_id) | Q(adncenate=area_id),
            ainfecegre__isnull=True,
            ainestado=1
        ).select_related('genpacien')

    results = []
    
    # Obtener códigos permitidos del Maestro Técnico (Medicamentos Estériles)
    allowed_codes = set(MedicamentoEsteril.objects.values_list('cod_med_1', flat=True)) | \
                    set(MedicamentoEsteril.objects.values_list('cod_med_2', flat=True))
    allowed_codes.discard(None)
    allowed_codes.discard('')
    allowed_codes.discard('N/A')

    for ing in ingresos:
        paciente = ing.genpacien
        if not paciente:
            continue
        
        # 2. SISTEMA DE BÚSQUEDA DE CAMA
        num_cama = 'S/C'
        cama = None
        if ing.hpndefcam:
            cama = Hpndefcam.objects.using('readonly').filter(oid=ing.hpndefcam).first()
        if not cama:
            estancia = Hpnestanc.objects.using('readonly').filter(adningres=ing.oid, hesfecsal__isnull=True).order_by('-hesfecing').first()
            if estancia and estancia.hpndefcam:
                cama = Hpndefcam.objects.using('readonly').filter(oid=estancia.hpndefcam).first()
        
        if cama:
            num_cama = cama.hcacodigo or cama.hcanumhabi or cama.hcanombre
            if any(x in str(num_cama).upper() for x in ['URGENCIAS', 'OBSERVACION', 'PISO']):
                if cama.hcacodigo and cama.hcacodigo.strip():
                    num_cama = cama.hcacodigo
        
        # 3. Obtener la última fórmula médica (el folio más reciente que SÍ tenga medicamentos)
        folios_ingreso = Hcnfolio.objects.using('readonly').filter(
            adningreso=ing.oid
        ).order_by('-hcfecfol').values_list('oid', flat=True)
        
        folio = None
        formulas = None
        
        for f_oid in folios_ingreso:
            formulas_folio = Hcnmedpac.objects.using('readonly').filter(
                hcnfolio=f_oid
            ).filter(
                Q(hcsfecsus__isnull=True) | Q(hcsfecsus__gt=timezone.now())
            )
            
            if formulas_folio.exists():
                folio = Hcnfolio.objects.using('readonly').get(oid=f_oid)
                formulas = formulas_folio
                break
                
        if not folio or not formulas:
            continue
            
        # 4. Procesar solo los medicamentos de esa última fórmula
        meds_data = []

        for f in formulas:
            try:
                producto = Innproduc.objects.using('readonly').get(oid=f.innproduc)
                
                # FILTRO PRINCIPAL ELIMINADO: Ahora muestra TODOS los medicamentos
                # (no solo los del Maestro Técnico de Estériles)
                # Filtro inteligente de medicamentos vs insumos
                nombre_prod = (producto.iprdescor or '').upper()
                
                # 1. Lista negra de términos que definen insumos clínicos
                insumo_keywords = [
                    'JERINGA', 'GUANTE', 'GASA', 'CATETER', 'EQUIPO', 'SONDA', 
                    'MICROGOTERO', 'MACROGOTERO', 'ALGODON', 'ALCOHOL', 'TAPABOCA',
                    'MASCARILLA', 'ELECTRODO', 'BOLSA', 'TUBO', 'AGUJA', 'CANULA',
                    'JABON', 'RECOLECTOR', 'VENDA', 'ESPARADRAPO', 'MICROPORE'
                ]
                
                # 2. Si el nombre contiene un insumo, lo saltamos
                if any(k in nombre_prod for k in insumo_keywords):
                    continue
                    
                # 3. Validación de dosis/frecuencia: Los medicamentos usualmente tienen estos campos
                if not f.hcsdosis and not f.hcsfrecmed:
                    # Si no tiene dosis ni frecuencia, es casi seguro que es un suministro
                    continue

                # Intentar obtener la vía desde los campos descriptivos de la fórmula
                via = (f.hcsviaadmntp or f.hcsviadforotr or "").strip()
                if not via or via.upper() == "N/D":
                    via = "N/D"

                # Obtener la forma farmacéutica desde el producto (más específico que el grupo farmacológico)
                forma_farmaceutica = (producto.iprforfar or "").strip()
                
                # Si no tiene forma farmacéutica, intentamos con el grupo farmacológico (Innfarcol)
                if not forma_farmaceutica:
                    forma_farmaceutica = "N/D"
                    if producto.innfarcol:
                        try:
                            farcol = Innfarcol.objects.using('readonly').get(oid=producto.innfarcol)
                            forma_farmaceutica = (farcol.farnombre or "N/D").strip()
                        except:
                            pass

                # Si la vía sigue siendo N/D, intentamos resolverla por convenciones o heurística
                if via == "N/D":
                    # 1. Buscar en tabla de convenciones por forma farmacéutica
                    if forma_farmaceutica != "N/D":
                        convencion = ConvencionFormaFarmaceutica.objects.filter(
                            forma_farmaceutica__iexact=forma_farmaceutica
                        ).first()
                        if convencion:
                            via = convencion.via

                    # 2. Si sigue siendo N/D, aplicar mapeo por códigos enteros conocidos
                    if via == "N/D":
                        if f.hcsviaadm == 0:
                            via = "ORAL"
                        elif f.hcsviaadm in [27, 28, 29, 128]:
                            via = "IV"
                        elif f.hcsviaadm == 137:
                            via = "SC"

                    # 3. Heurística basada en el nombre de la forma o el producto
                    if via == "N/D":
                        desc_upper = (producto.iprdescor or "").upper()
                        forma_upper = forma_farmaceutica.upper()
                        
                        if "TABLETA" in forma_upper or "CAPSULA" in forma_upper or "ORAL" in forma_upper or "TABLETA" in desc_upper:
                            via = "ORAL"
                        elif "INYECTABLE" in forma_upper or "AMPOLLA" in forma_upper or "INYECTABLE" in desc_upper:
                            via = "IV"

                meds_data.append({
                    'codigo': producto.iprcodigo,
                    'nombre': f"{producto.iprdescor} ({(forma_farmaceutica if forma_farmaceutica != 'N/D' else 'Medicamento')})",
                    'dosis': f"{f.hcsdosis or ''} {f.hcsfrecmed or ''}h",
                    'cantidad': str(f.hcscanti or '0'),
                    'via': via
                })
            except Exception:
                continue

        # Calcular edad y aseguradora (fuera del if meds_data para que siempre estén disponibles)
        try:
            today = datetime.now()
            born = paciente.gpafecnac
            if born:
                edad = today.year - born.year - ((today.month, today.day) < (born.month, born.day))
            else:
                edad = "N/A"
        except:
            edad = "N/A"

        aseguradora = "Particular"
        if ing.gendetcon:
            try:
                aseguradora = ing.gendetcon.gdenombre
            except:
                pass

        results.append({
            'paciente_id': paciente.pacnumdoc,
            'paciente_nombre': f"{paciente.pacprinom or ''} {paciente.pacsegnom or ''} {paciente.pacpriape or ''} {paciente.pacsegape or ''}".strip(),
            'fecha_nacimiento': paciente.gpafecnac.strftime('%d/%m/%Y') if paciente.gpafecnac else 'N/A',
            'edad': str(edad),
            'aseguradora': aseguradora,
            'cama': num_cama,
            'folio': folio.hcnumfol if folio else "N/D",
            'medicamentos': meds_data,
            'registros': len(meds_data)
        })

    return JsonResponse({'results': results})

# --- PACIENTES ACTIVOS EXCLUSIVO NEONATOS ---

class NeonatosPacientesActivosView(LoginRequiredMixin, TemplateView):
    template_name = 'CentralDeMezclas/neonatos/pacientes_activos.html'

@login_required
def api_formulas_neonatos(request):
    """
    API exclusiva para Neonatos: retorna pacientes activos SOLO del área neonatal,
    con filtro opcional por hora de prescripción.
    """
    from consultas_externas.models import Hpnsubgru
    
    hora_desde = request.GET.get('hora_desde', '')
    
    # 1. Identificar las salas/subgrupos que contienen "NEONAT"
    neo_subgrupos = Hpnsubgru.objects.using('readonly').filter(
        Q(hsunombre__icontains='neonat')
    ).values_list('oid', flat=True)
    
    # 2. Obtener camas que pertenezcan a esos subgrupos
    neo_camas = Hpndefcam.objects.using('readonly').filter(
        hpnsubgru__in=neo_subgrupos
    ).values_list('oid', flat=True)
    
    # 3. Obtener estancias activas en esas camas
    neo_estancias = Hpnestanc.objects.using('readonly').filter(
        hpndefcam__in=neo_camas,
        hesfecsal__isnull=True
    )
    
    adm_oids = [s.adningres for s in neo_estancias if s.adningres]
    
    # 4. Obtener los ingresos activos
    ingresos = Adningreso.objects.using('readonly').filter(
        oid__in=adm_oids,
        ainfecegre__isnull=True
    ).select_related('genpacien')
    
    results = []
    
    for ing in ingresos:
        paciente = ing.genpacien
        if not paciente:
            continue
        
        # Buscar cama
        num_cama = 'S/C'
        cama = None
        if ing.hpndefcam:
            cama = Hpndefcam.objects.using('readonly').filter(oid=ing.hpndefcam).first()
        if not cama:
            estancia = Hpnestanc.objects.using('readonly').filter(adningres=ing.oid, hesfecsal__isnull=True).order_by('-hesfecing').first()
            if estancia and estancia.hpndefcam:
                cama = Hpndefcam.objects.using('readonly').filter(oid=estancia.hpndefcam).first()
        
        if cama:
            num_cama = cama.hcacodigo or cama.hcanumhabi or cama.hcanombre
        
        # Obtener la última fórmula médica
        folios_ingreso = Hcnfolio.objects.using('readonly').filter(
            adningreso=ing.oid
        ).order_by('-hcfecfol').values_list('oid', flat=True)
        
        folio = None
        formulas = None
        
        for f_oid in folios_ingreso:
            formulas_folio = Hcnmedpac.objects.using('readonly').filter(
                hcnfolio=f_oid
            ).filter(
                Q(hcsfecsus__isnull=True) | Q(hcsfecsus__gt=timezone.now())
            )
            
            # Filtro por hora si se especificó
            if hora_desde:
                try:
                    hora_h = int(hora_desde)
                    from datetime import datetime as dt_local
                    hoy = dt_local.now().replace(hour=hora_h, minute=0, second=0, microsecond=0)
                    formulas_folio = formulas_folio.filter(hcsfecmed__gte=hoy)
                except (ValueError, TypeError):
                    pass
            
            if formulas_folio.exists():
                folio = Hcnfolio.objects.using('readonly').get(oid=f_oid)
                formulas = formulas_folio
                break
        
        if not folio or not formulas:
            continue
        
        # Procesar medicamentos
        meds_data = []
        insumo_keywords = [
            'JERINGA', 'GUANTE', 'GASA', 'CATETER', 'EQUIPO', 'SONDA',
            'MICROGOTERO', 'MACROGOTERO', 'ALGODON', 'ALCOHOL', 'TAPABOCA',
            'MASCARILLA', 'ELECTRODO', 'BOLSA', 'TUBO', 'AGUJA', 'CANULA',
            'JABON', 'RECOLECTOR', 'VENDA', 'ESPARADRAPO', 'MICROPORE'
        ]
        
        for f in formulas:
            try:
                producto = Innproduc.objects.using('readonly').get(oid=f.innproduc)
                nombre_prod = (producto.iprdescor or '').upper()
                
                if any(k in nombre_prod for k in insumo_keywords):
                    continue
                if not f.hcsdosis and not f.hcsfrecmed:
                    continue
                
                via = (f.hcsviaadmntp or f.hcsviadforotr or "").strip()
                if not via or via.upper() == "N/D":
                    via = "N/D"
                
                forma_farmaceutica = (producto.iprforfar or "").strip()
                if not forma_farmaceutica:
                    forma_farmaceutica = "N/D"
                    if producto.innfarcol:
                        try:
                            farcol = Innfarcol.objects.using('readonly').get(oid=producto.innfarcol)
                            forma_farmaceutica = (farcol.farnombre or "N/D").strip()
                        except:
                            pass
                
                if via == "N/D":
                    if forma_farmaceutica != "N/D":
                        convencion = ConvencionFormaFarmaceutica.objects.filter(
                            forma_farmaceutica__iexact=forma_farmaceutica
                        ).first()
                        if convencion:
                            via = convencion.via
                    if via == "N/D":
                        if f.hcsviaadm == 0:
                            via = "ORAL"
                        elif f.hcsviaadm in [27, 28, 29, 128]:
                            via = "IV"
                        elif f.hcsviaadm == 137:
                            via = "SC"
                    if via == "N/D":
                        desc_upper = (producto.iprdescor or "").upper()
                        forma_upper = forma_farmaceutica.upper()
                        if "TABLETA" in forma_upper or "CAPSULA" in forma_upper or "ORAL" in forma_upper:
                            via = "ORAL"
                        elif "INYECTABLE" in forma_upper or "AMPOLLA" in forma_upper:
                            via = "IV"
                
                meds_data.append({
                    'codigo': producto.iprcodigo,
                    'nombre': f"{producto.iprdescor} ({(forma_farmaceutica if forma_farmaceutica != 'N/D' else 'Medicamento')})",
                    'dosis': f"{f.hcsdosis or ''} {f.hcsfrecmed or ''}h",
                    'cantidad': str(f.hcscanti or '0'),
                    'via': via
                })
            except Exception:
                continue
        
        # Calcular edad y filtrar por seguridad (solo menores de 2 años en Neonatos)
        try:
            today = timezone.now()
            born = paciente.gpafecnac
            if born:
                # Convertir born a objeto aware si es naive para comparar con timezone.now()
                # O usar datetime simple si la DB es simple
                from datetime import datetime as dt_plain
                today_plain = dt_plain.now()
                born_plain = born.replace(tzinfo=None) if hasattr(born, 'tzinfo') else born
                edad = today_plain.year - born_plain.year - ((today_plain.month, today_plain.day) < (born_plain.month, born_plain.day))
                
                # FILTRO DE SEGURIDAD: En neonatos no deberían haber niños de más de 2 años
                if edad > 2:
                    continue
            else:
                edad = "N/A"
        except Exception as e:
            print(f"Error calculando edad: {e}")
            edad = "N/A"
        
        aseguradora = "Particular"
        if ing.gendetcon:
            try:
                aseguradora = ing.gendetcon.gdenombre
            except:
                pass
        
        results.append({
            'paciente_id': paciente.pacnumdoc,
            'paciente_nombre': f"{paciente.pacprinom or ''} {paciente.pacsegnom or ''} {paciente.pacpriape or ''} {paciente.pacsegape or ''}".strip(),
            'fecha_nacimiento': paciente.gpafecnac.strftime('%d/%m/%Y') if paciente.gpafecnac else 'N/A',
            'edad': str(edad),
            'aseguradora': aseguradora,
            'cama': num_cama,
            'folio': folio.hcnumfol if folio else "N/D",
            'medicamentos': meds_data,
            'registros': len(meds_data),
            'ingreso_id': ing.oid
        })
    
    return JsonResponse({'results': results, 'count': len(results)})

# --- FUNCIONES DE EXPORTACIÓN (XLS/PDF) ---

def export_reempaque_xls(request):
    """Exportar Catálogo FRFAR-089 inyectando datos en la plantilla CENTRAL_TOTAL (D:\\central de mezclas)"""
    template_path = r'D:\central de mezclas\CENTRAL_TOTAL.xlsx'
    
    try:
        wb = openpyxl.load_workbook(template_path)
        # Intentamos buscar una hoja que se llame FRFAR-089 o similar, si no usamos la activa
        ws = wb.active
        for sheet in wb.sheetnames:
            if "089" in sheet:
                ws = wb[sheet]
                break
    except Exception as e:
        return HttpResponse(f"Error: No se encontró la plantilla en {template_path} o está bloqueada. <br> Detalle: {e}", status=404)

    # El usuario solicita iniciar en la casilla A26 (Fila 26)
    objs = ReempaqueMedicamento.objects.all().order_by('nombre')
    
    start_row = 26
    current_row = start_row
    
    for obj in objs:
        ws.cell(row=current_row, column=1, value=obj.nombre)
        ws.cell(row=current_row, column=2, value=obj.concentracion)
        ws.cell(row=current_row, column=3, value=obj.convencion.forma_farmaceutica if obj.convencion else "N/A")
        ws.cell(row=current_row, column=4, value=obj.convencion.via if obj.convencion else "N/A")
        ws.cell(row=current_row, column=5, value=obj.alerta.codigo if obj.alerta else "N/A")
        ws.cell(row=current_row, column=6, value=obj.condiciones_almacenamiento)
        ws.cell(row=current_row, column=7, value=obj.fotosensibilidad)
        current_row += 1
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=CENTRAL_TOTAL_FRFAR-089.xlsx'
    return response

def export_matriz_xls(request):
    """Exportar Matriz FRFAR-090 a Excel con estructura solicitada"""
    objs = ReempaqueOrden.objects.all().order_by('-fecha_produccion')
    data = []
    for o in objs:
        data.append({
            'Lote Interno': o.lote_interno,
            'Medicamento': o.medicamento.nombre,
            'Concentración': o.medicamento.concentracion,
            'Forma Farmacéutica': o.medicamento.convencion.forma_farmaceutica if o.medicamento.convencion else "N/A",
            'Laboratorio': o.laboratorio or '-',
            'Lote': o.lote_fabricante,
            'Fecha Vencimiento': o.fecha_vencimiento.strftime('%y-%m') if o.fecha_vencimiento else 'N/A',
            'Registro INVIMA': o.registro_invima or '-',
            'Cantidad': o.cantidad_a_reempacar,
            'Temperatura': o.medicamento.condiciones_almacenamiento,
            'Fotosensible': o.medicamento.fotosensibilidad,
            'SEMAFORIZACIÓN': o.medicamento.alerta.codigo if o.medicamento.alerta else (o.semaforizacion or '-'),
            'Estado': o.get_estado_display()
        })
    
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='FRFAR-090')
    
    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=FRFAR-090_Matriz.xlsx'
    return response

def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = io.BytesIO()
    # xhtml2pdf necesita un stream binario
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None

def export_matriz_pdf(request):
    """Exportar Matriz FRFAR-090 a PDF institucional"""
    ordenes = ReempaqueOrden.objects.all().order_by('-fecha_produccion')
    context = {
        'ordenes': ordenes,
        'fecha': pd.Timestamp.now().strftime('%Y-%m-%d'),
        'formato': 'FRFAR-090'
    }
    response = render_to_pdf('CentralDeMezclas/reempaque/pdf_matriz.html', context)
    if response:
        response['Content-Disposition'] = 'attachment; filename="FRFAR-090_Matriz.pdf"'
        return response
    return HttpResponse("Error generando PDF", status=500)

# --- VISTAS DE CONVENCIONES Y ALERTAS ---

class ConvencionListView(LoginRequiredMixin, SearchFilterMixin, ListView):
    model = ConvencionFormaFarmaceutica
    template_name = 'CentralDeMezclas/config/convencion_list.html'
    context_object_name = 'convenciones'

class ConvencionCreateView(LoginRequiredMixin, CreateView):
    model = ConvencionFormaFarmaceutica
    template_name = 'CentralDeMezclas/config/convencion_form.html'
    fields = '__all__'
    success_url = reverse_lazy('mezclas:convencion_list')

class ConvencionUpdateView(LoginRequiredMixin, UpdateView):
    model = ConvencionFormaFarmaceutica
    template_name = 'CentralDeMezclas/config/convencion_form.html'
    fields = '__all__'
    success_url = reverse_lazy('mezclas:convencion_list')

class ConvencionDeleteView(LoginRequiredMixin, DeleteView):
    model = ConvencionFormaFarmaceutica
    template_name = 'CentralDeMezclas/confirm_delete.html'
    success_url = reverse_lazy('mezclas:convencion_list')

class AlertaListView(LoginRequiredMixin, ListView):
    model = Alerta
    template_name = 'CentralDeMezclas/config/alerta_list.html'
    context_object_name = 'alertas'

class AlertaCreateView(LoginRequiredMixin, CreateView):
    model = Alerta
    template_name = 'CentralDeMezclas/config/alerta_form.html'
    fields = '__all__'
    success_url = reverse_lazy('mezclas:alerta_list')

class AlertaUpdateView(LoginRequiredMixin, UpdateView):
    model = Alerta
    template_name = 'CentralDeMezclas/config/alerta_form.html'
    fields = '__all__'
    success_url = reverse_lazy('mezclas:alerta_list')

class AlertaDeleteView(LoginRequiredMixin, DeleteView):
    model = Alerta
    template_name = 'CentralDeMezclas/confirm_delete.html'
    success_url = reverse_lazy('mezclas:alerta_list')


# --- VISTAS MEDICAMENTOS ONCOLÓGICOS ALTA COMPLEJIDAD (FRFAR-126) ---

class MedicamentoOncologicoListView(LoginRequiredMixin, ListView):
    model = MedicamentoOncologico
    template_name = 'CentralDeMezclas/oncologicos/medicamento_list.html'
    context_object_name = 'medicamentos'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total'] = MedicamentoOncologico.objects.count()
        return context

class MedicamentoOncologicoCreateView(LoginRequiredMixin, CreateView):
    model = MedicamentoOncologico
    template_name = 'CentralDeMezclas/oncologicos/medicamento_form.html'
    fields = '__all__'
    success_url = reverse_lazy('mezclas:oncologicos_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Nuevo Medicamento Oncológico'
        return context

class MedicamentoOncologicoUpdateView(LoginRequiredMixin, UpdateView):
    model = MedicamentoOncologico
    template_name = 'CentralDeMezclas/oncologicos/medicamento_form.html'
    fields = '__all__'
    success_url = reverse_lazy('mezclas:oncologicos_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Editar: {self.object.medicamento}'
        return context

class MedicamentoOncologicoDeleteView(LoginRequiredMixin, DeleteView):
    model = MedicamentoOncologico
    template_name = 'CentralDeMezclas/confirm_delete.html'
    success_url = reverse_lazy('mezclas:oncologicos_list')

@login_required
def export_oncologicos_xls(request):
    """Exportar catálogo de medicamentos oncológicos FRFAR-126 a Excel"""
    objs = MedicamentoOncologico.objects.all().order_by('cod')
    data = []
    for o in objs:
        data.append({
            'COD.': o.cod,
            'PRODUCTO': o.producto,
            'MEDICAMENTO': o.medicamento,
            'CONCENTRACION (mg, UI)': o.concentracion,
            'FORMA FARMACEUTICA': o.forma_farmaceutica,
            'VOLUMEN DE RECONSTITUCION': o.volumen_reconstitucion,
            'SOLUCION DE RECONSTITUCION': o.solucion_reconstitucion,
            'COD. DIL.': o.cod_dil,
            'ADMINISTRACION': o.administracion,
            'VOL. FINAL': o.vol_final,
            'VEHICULO': o.vehiculo,
            'Jeringa': o.jeringa,
            'Aguja': o.aguja,
            'ALMACENAMIENTO': o.almacenamiento,
            'OBSERVACIONES': o.proteccion_luz,
            'FECHA DE EXPIRACION': o.fecha_expiracion,
        })

    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='FRFAR-126')

    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=FRFAR-126_Oncologicos.xlsx'
    return response


# ========== PANEL ONCOLÓGICOS ==========

class OncologicoPanelView(LoginRequiredMixin, TemplateView):
    template_name = 'CentralDeMezclas/oncologicos/panel.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['stats'] = {
            'medicamentos': MedicamentoOncologico.objects.count(),
            'matrices': OncologicoMatriz.objects.count(),
            'ordenes': OncologicoOrdenProduccion.objects.count(),
            'alistamientos': OncologicoAlistamiento.objects.count(),
        }
        return context


# ========== FRFAR-127 MATRIZ ONCOLÓGICA ==========

class OncologicoMatrizListView(LoginRequiredMixin, ListView):
    model = OncologicoMatriz
    template_name = 'CentralDeMezclas/oncologicos/matriz_list.html'
    context_object_name = 'matrices'

class OncologicoMatrizCreateView(LoginRequiredMixin, CreateView):
    model = OncologicoMatriz
    template_name = 'CentralDeMezclas/oncologicos/medicamento_form.html'
    fields = '__all__'
    success_url = reverse_lazy('mezclas:onco_matriz_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Nueva Matriz Oncológica'
        return context

class OncologicoMatrizDetailView(LoginRequiredMixin, DetailView):
    model = OncologicoMatriz
    template_name = 'CentralDeMezclas/oncologicos/matriz_detalle.html'
    context_object_name = 'matriz'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['items'] = self.object.items.all()
        return context

class OncologicoMatrizDeleteView(LoginRequiredMixin, DeleteView):
    model = OncologicoMatriz
    template_name = 'CentralDeMezclas/confirm_delete.html'
    success_url = reverse_lazy('mezclas:onco_matriz_list')

class OncologicoMatrizItemCreateView(LoginRequiredMixin, CreateView):
    model = OncologicoMatrizItem
    template_name = 'CentralDeMezclas/oncologicos/medicamento_form.html'
    fields = ['lote_interno', 'paciente_nombre', 'identificacion', 'cama', 'servicio',
              'medicamento_base', 'cod', 'medicamento', 'concentracion', 'forma_farmaceutica',
              'dosis', 'frecuencia', 'volumen_final', 'lote', 'fecha_vencimiento',
              'solucion_diluyente', 'viales_ampollas', 'vol_dilucion', 'vol_dosis',
              'vol_final_unidosis', 'via_admon']

    def get_initial(self):
        initial = super().get_initial()
        initial['matriz'] = self.kwargs['matriz_pk']
        return initial

    def form_valid(self, form):
        form.instance.matriz = get_object_or_404(OncologicoMatriz, pk=self.kwargs['matriz_pk'])
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('mezclas:onco_matriz_detalle', kwargs={'pk': self.kwargs['matriz_pk']})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Agregar Ítem a la Matriz'
        return context

class OncologicoMatrizItemDeleteView(LoginRequiredMixin, DeleteView):
    model = OncologicoMatrizItem
    template_name = 'CentralDeMezclas/confirm_delete.html'

    def get_success_url(self):
        return reverse_lazy('mezclas:onco_matriz_detalle', kwargs={'pk': self.object.matriz.pk})


# ========== FRFAR-178 ORDEN DE PRODUCCIÓN ONCOLÓGICA ==========

class OncologicoOrdenListView(LoginRequiredMixin, ListView):
    model = OncologicoOrdenProduccion
    template_name = 'CentralDeMezclas/oncologicos/orden_list.html'
    context_object_name = 'ordenes'

class OncologicoOrdenCreateView(LoginRequiredMixin, CreateView):
    model = OncologicoOrdenProduccion
    template_name = 'CentralDeMezclas/oncologicos/medicamento_form.html'
    fields = '__all__'
    success_url = reverse_lazy('mezclas:onco_orden_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Nueva Orden de Producción Oncológica'
        return context

class OncologicoOrdenDetailView(LoginRequiredMixin, DetailView):
    model = OncologicoOrdenProduccion
    template_name = 'CentralDeMezclas/oncologicos/orden_detalle.html'
    context_object_name = 'orden'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['items'] = self.object.items.all()
        return context

class OncologicoOrdenDeleteView(LoginRequiredMixin, DeleteView):
    model = OncologicoOrdenProduccion
    template_name = 'CentralDeMezclas/confirm_delete.html'
    success_url = reverse_lazy('mezclas:onco_orden_list')

class OncologicoOrdenItemCreateView(LoginRequiredMixin, CreateView):
    model = OncologicoOrdenItem
    template_name = 'CentralDeMezclas/oncologicos/medicamento_form.html'
    fields = ['paciente_nombre', 'identificacion', 'cama', 'medicamento', 'via_administracion',
              'viales_ampollas', 'vehiculo_reconstitucion', 'volumen_reconstitucion',
              'dosis', 'vol_dosis', 'vehiculo_unidosis', 'vol_final', 'concentracion_final',
              'lote_interno', 'cantidad', 'recibido']

    def form_valid(self, form):
        form.instance.orden = get_object_or_404(OncologicoOrdenProduccion, pk=self.kwargs['orden_pk'])
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('mezclas:onco_orden_detalle', kwargs={'pk': self.kwargs['orden_pk']})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Agregar Ítem a la Orden de Producción'
        return context

class OncologicoOrdenItemDeleteView(LoginRequiredMixin, DeleteView):
    model = OncologicoOrdenItem
    template_name = 'CentralDeMezclas/confirm_delete.html'

    def get_success_url(self):
        return reverse_lazy('mezclas:onco_orden_detalle', kwargs={'pk': self.object.orden.pk})


# ========== FRFAR-162 ALISTAMIENTO ONCOLÓGICO ==========

class OncologicoAlistamientoListView(LoginRequiredMixin, ListView):
    model = OncologicoAlistamiento
    template_name = 'CentralDeMezclas/oncologicos/alistamiento_list.html'
    context_object_name = 'alistamientos'

class OncologicoAlistamientoCreateView(LoginRequiredMixin, CreateView):
    model = OncologicoAlistamiento
    template_name = 'CentralDeMezclas/oncologicos/medicamento_form.html'
    fields = '__all__'
    success_url = reverse_lazy('mezclas:onco_alistamiento_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Nuevo Alistamiento y Conciliación Oncológico'
        return context

class OncologicoAlistamientoDetailView(LoginRequiredMixin, DetailView):
    model = OncologicoAlistamiento
    template_name = 'CentralDeMezclas/oncologicos/alistamiento_detalle.html'
    context_object_name = 'alistamiento'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['items'] = self.object.items.all()
        return context

class OncologicoAlistamientoDeleteView(LoginRequiredMixin, DeleteView):
    model = OncologicoAlistamiento
    template_name = 'CentralDeMezclas/confirm_delete.html'
    success_url = reverse_lazy('mezclas:onco_alistamiento_list')

class OncologicoAlistamientoItemCreateView(LoginRequiredMixin, CreateView):
    model = OncologicoAlistamientoItem
    template_name = 'CentralDeMezclas/oncologicos/medicamento_form.html'
    fields = ['material', 'lote_fabricante', 'fecha_vencimiento', 'cantidad_solicitada',
              'cantidad_ingresada', 'aprovechamiento', 'producto_terminado']

    def form_valid(self, form):
        form.instance.alistamiento = get_object_or_404(OncologicoAlistamiento, pk=self.kwargs['alistamiento_pk'])
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('mezclas:onco_alistamiento_detalle', kwargs={'pk': self.kwargs['alistamiento_pk']})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Agregar Material al Alistamiento'
        return context

class OncologicoAlistamientoItemDeleteView(LoginRequiredMixin, DeleteView):
    model = OncologicoAlistamientoItem
    template_name = 'CentralDeMezclas/confirm_delete.html'

    def get_success_url(self):
        return reverse_lazy('mezclas:onco_alistamiento_detalle', kwargs={'pk': self.object.alistamiento.pk})
# --- VISTAS MÓDULO NEONATOS (NEOHUDN1) ---

class NeonatosPanelView(LoginRequiredMixin, TemplateView):
    template_name = 'CentralDeMezclas/neonatos/panel.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['stats'] = {
            'medicamentos': NeonatosMedicamento.objects.count(),
            'matrices': NeonatosMatriz.objects.count(),
            'ordenes': NeonatosOrdenProduccion.objects.count(),
            'alistamientos': NeonatosAlistamiento.objects.count(),
        }
        return context

class NeonatosMatrizForm(forms.ModelForm):
    # Usamos CharField para capturar los IDs (Cédulas) del HUDN via Select2
    # Esto evita fallos de validación de ForeignKey antes de sincronizar en form_valid
    quien_prepara = forms.CharField(required=False, label="Quien Prepara")
    preelaboracion = forms.CharField(required=False, label="Preelaboración")
    alistamiento = forms.CharField(required=False, label="Alistamiento")
    quien_cobra = forms.CharField(required=False, label="Quien Cobra")
    jefe_produccion = forms.CharField(required=False, label="Jefe de Producción")
    jefe_control_calidad = forms.CharField(required=False, label="Jefe de C. Calidad")
    digitador = forms.CharField(required=False, label="Digitador")

    class Meta:
        model = NeonatosMatriz
        fields = ['fecha', 'orden_produccion', 'quien_prepara', 'preelaboracion', 'alistamiento', 'quien_cobra', 'jefe_produccion', 'jefe_control_calidad', 'digitador']

class NeonatosListView(LoginRequiredMixin, SearchFilterMixin, ListView):
    model = NeonatosMatriz
    template_name = 'CentralDeMezclas/neonatos/list.html'
    context_object_name = 'matrices'
    ordering = ['-fecha', '-pk']

class NeonatosCreateView(LoginRequiredMixin, CreateView):
    model = NeonatosMatriz
    form_class = NeonatosMatrizForm
    template_name = 'CentralDeMezclas/neonatos/form.html'
    success_url = reverse_lazy('mezclas:neonatos_list')

    def get_initial(self):
        initial = super().get_initial()
        now = timezone.now()
        count = NeonatosMatriz.objects.filter(fecha=now.date()).count() + 1
        initial['orden_produccion'] = f"NEO-{now.strftime('%y%m%d')}-{count:03d}"
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Nueva Matriz de Neonatos'
        return context

    def form_valid(self, form):
        # 1. Procesar Roles (Sincronizar con HUDN si es necesario)
        roles_fields = ['quien_prepara', 'preelaboracion', 'alistamiento', 'quien_cobra', 'jefe_produccion', 'jefe_control_calidad', 'digitador']
        
        for field in roles_fields:
            cedula = self.request.POST.get(field)
            if cedula:
                # Buscar o crear funcionario localmente basándose en la cédula del HUDN
                funcionario, created = Funcionario.objects.get_or_create(
                    cedula=cedula,
                    defaults={'nombre_completo': 'PENDIENTE DE HUDN'}
                )
                
                if created or funcionario.nombre_completo == 'PENDIENTE DE HUDN':
                    # Actualizar con datos reales de Gentercer
                    tercero = Gentercer.objects.using('readonly').filter(ternumdoc=cedula).first()
                    if tercero:
                        funcionario.nombre_completo = tercero.ternomcom or f"{tercero.terprinom} {tercero.terpriape}".strip()
                        funcionario.save()
                
                setattr(form.instance, field, funcionario)

        # 2. Guardar el encabezado
        response = super().form_valid(form)
        matriz = self.object

        # 3. Procesar datos de la cuadrícula (matriz_data)
        matriz_data_json = self.request.POST.get('matriz_data')
        if matriz_data_json:
            try:
                items_data = json.loads(matriz_data_json)
                for i, row in enumerate(items_data):
                    # row format based on columns in form.html:
                    # [0: CC, 1: Nombre, 2: Cama, 3: Servicio, 4: Medicamento, 5: Dosis, 6: Vol Final, 7: Dilucion, 8: Lote]
                    if row[0] and row[1]: # Validar ID y Nombre
                        NeonatosMatrizItem.objects.create(
                            matriz=matriz,
                            lote_interno=f"{matriz.orden_produccion}-{i+1:02d}",
                            identificacion=row[0],
                            paciente_nombre=row[1],
                            cama=row[2] or '',
                            servicio=row[3] or 'NEONATOS',
                            medicamento=row[4] or '',
                            dosis=row[5] or '',
                            volumen_final=row[6] or '',
                            solucion_diluyente=row[7] or '',
                            lote=row[8] or ''
                        )
            except Exception as e:
                print(f"Error procesando matriz_data: {e}")

        return response

# --- VISTAS CATÁLOGO NEONATOS ---

class NeonatosMedicamentoListView(LoginRequiredMixin, ListView):
    model = NeonatosMedicamento
    template_name = 'CentralDeMezclas/neonatos/medicamento_list.html'
    context_object_name = 'medicamentos'

class NeonatosMedicamentoCreateView(LoginRequiredMixin, CreateView):
    model = NeonatosMedicamento
    template_name = 'CentralDeMezclas/neonatos/medicamento_form.html'
    fields = '__all__'
    success_url = reverse_lazy('mezclas:neonatos_med_list')

class NeonatosMedicamentoUpdateView(LoginRequiredMixin, UpdateView):
    model = NeonatosMedicamento
    template_name = 'CentralDeMezclas/neonatos/medicamento_form.html'
    fields = '__all__'
    success_url = reverse_lazy('mezclas:neonatos_med_list')

class NeonatosMedicamentoDeleteView(LoginRequiredMixin, DeleteView):
    model = NeonatosMedicamento
    template_name = 'CentralDeMezclas/confirm_delete.html'
    success_url = reverse_lazy('mezclas:neonatos_med_list')

# --- VISTAS MÓDULO MAGISTRALES (MAGHUDN1) ---

# --- VISTAS UNIDOSIS PRODUCCIÓN (MAGHUDN1) ---

class UnidosisProduccionListView(LoginRequiredMixin, ListView):
    model = UnidosisProduccionOrden
    template_name = 'CentralDeMezclas/unidosis_prod/list.html'
    context_object_name = 'ordenes'

class UnidosisProduccionCreateView(LoginRequiredMixin, CreateView):
    model = UnidosisProduccionOrden
    template_name = 'CentralDeMezclas/unidosis_prod/form.html'
    fields = '__all__'
    success_url = reverse_lazy('mezclas:unidosis_prod_list')

# --- VISTAS MÓDULO NUTRICIÓN PARENTERAL (NPTHUDN1) ---

class NptPanelView(LoginRequiredMixin, TemplateView):
    template_name = 'CentralDeMezclas/nutricion/panel.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['stats'] = {
            'matrices': NptMatriz.objects.count(),
            'ordenes': NptOrdenProduccion.objects.count(),
            'alistamientos': NptAlistamiento.objects.count(),
        }
        return context

class NptListView(LoginRequiredMixin, ListView):
    model = NptMatriz
    template_name = 'CentralDeMezclas/nutricion/list.html'
    context_object_name = 'matrices'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Matrices de Nutrición Parenteral'
        return context

class NptCreateView(LoginRequiredMixin, CreateView):
    model = NptMatriz
    template_name = 'CentralDeMezclas/nutricion/form.html'
    fields = '__all__'
    success_url = reverse_lazy('mezclas:nutricion_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Nueva Matriz NPT'
        return context


# --- PROCESOS DIARIOS E INFORMES (REPORTES) ---

class ProcesosDiariosView(LoginRequiredMixin, TemplateView):
    template_name = 'CentralDeMezclas/reportes/procesos_diarios.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        fecha_str = self.request.GET.get('fecha', timezone.now().date().isoformat())
        try:
            fecha = timezone.datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            fecha = timezone.now().date()
            
        context['fecha_consulta'] = fecha
        
        # Consolidación de Actividad por Módulo
        context['resumen_diario'] = {
            'esteriles': UnidosisOrden.objects.filter(periodo__fecha=fecha).count(),
            'oncologicos': OncologicoMatrizItem.objects.filter(matriz__fecha=fecha).count(),
            'reempaque_reenvase': ReempaqueOrden.objects.filter(fecha_produccion__date=fecha).count(),
            'neonatos': NeonatosMatrizItem.objects.filter(matriz__fecha=fecha).count(),
            'unidosis_prod': UnidosisProduccionOrden.objects.filter(fecha=fecha).count(),
            'nutriciones': NptMatriz.objects.filter(fecha=fecha).count(),
        }
        
        # Detalle para las tablas
        context['esteriles_items'] = UnidosisOrden.objects.filter(periodo__fecha=fecha)
        context['oncologicos_items'] = OncologicoMatrizItem.objects.filter(matriz__fecha=fecha)
        context['unidosis_prod_items'] = UnidosisProduccionOrden.objects.filter(fecha=fecha)
        context['neonatos_items'] = NeonatosMatrizItem.objects.filter(matriz__fecha=fecha)
        
        return context

@login_required
def generar_informe_diario_pdf(request):
    """Generación de PDF consolidado de procesos diarios"""
    fecha_str = request.GET.get('fecha', timezone.now().date().isoformat())
    try:
        fecha = timezone.datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except ValueError:
        fecha = timezone.now().date()
    
    # Recopilar todos los datos del día
    datos = {
        'fecha': fecha,
        'esteriles': UnidosisOrden.objects.filter(periodo__fecha=fecha),
        'oncologicos': OncologicoMatrizItem.objects.filter(matriz__fecha=fecha),
        'neonatos': NeonatosMatrizItem.objects.filter(matriz__fecha=fecha),
        'unidosis_prod': UnidosisProduccionOrden.objects.filter(fecha=fecha),
        'nutriciones': NptMatriz.objects.filter(fecha=fecha),
        'reempaque': ReempaqueOrden.objects.filter(fecha_produccion__date=fecha),
    }
    
    response = render_to_pdf('CentralDeMezclas/reportes/informe_diario_pdf.html', datos)
    if response:
        response['Content-Disposition'] = f'attachment; filename="Informe_Diario_{fecha_str}.pdf"'
        return response
    return HttpResponse("Error generando informe PDF", status=500)



