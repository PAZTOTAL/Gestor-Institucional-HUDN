import io
from django.http import HttpResponse
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser
from . import PysAPIView as APIView

from django.db.models import Prefetch
from ..models import SolicitudPS, SolicitudPSArchivo, EncuestaRetiro
from ..permissions import require_roles

# Etiquetas legibles para los aspectos de la encuesta
_ASPECT_LABELS = {
    'calif_compañeros':     'Relación con compañeros',
    'calif_formacion':      'Formación',
    'calif_ambiente':       'Ambiente de trabajo',
    'calif_reconocimiento': 'Reconocimiento de la labor',
    'calif_carga_trabajo':  'Carga de trabajo',
    'calif_superior':       'Relación con el superior inmediato',
    'calif_beneficios':     'Beneficios sociales',
    'calif_salario':        'Salario',
    'calif_valores':        'Valores de la entidad',
    'calif_cultura':        'Cultura Organizacional',
    'calif_trabajo_equipo': 'Trabajo en equipo',
}

MAX_ARCHIVO_MB = 10
EXTS_PERMITIDAS = {'pdf', 'png', 'jpg', 'jpeg'}


class SolicitudPSCreateView(APIView):
    """Endpoint público — no requiere autenticación."""
    permission_classes = []
    authentication_classes = []
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        nombre = request.data.get('nombre', '').strip()
        identificacion = request.data.get('identificacion', '').strip()
        correo = request.data.get('correo', '').strip()
        coordinador = request.data.get('coordinador', '').strip()

        if not nombre or not identificacion:
            return Response(
                {'detail': 'Nombre e identificación son obligatorios'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not identificacion.isdigit():
            return Response(
                {'detail': 'La identificación debe contener solo números'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        archivos_data = []
        for archivo in request.FILES.getlist('archivos'):
            ext = archivo.name.rsplit('.', 1)[-1].lower() if '.' in archivo.name else ''
            if ext not in EXTS_PERMITIDAS:
                continue
            contenido = archivo.read()
            if len(contenido) > MAX_ARCHIVO_MB * 1024 * 1024:
                return Response(
                    {'detail': f'El archivo {archivo.name} supera {MAX_ARCHIVO_MB} MB'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            mime = archivo.content_type or ('application/pdf' if ext == 'pdf' else f'image/{ext}')
            archivos_data.append({
                'nombre_original': archivo.name,
                'contenido': contenido,
                'tipo_mime': mime,
            })

        if not archivos_data:
            return Response(
                {'detail': 'Adjunta al menos un archivo PDF o imagen válido'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        sol = SolicitudPS.objects.create(
            nombre=nombre,
            identificacion=identificacion,
            correo=correo or None,
            coordinador=coordinador or None,
        )
        for a in archivos_data:
            SolicitudPSArchivo.objects.create(
                solicitud=sol,
                nombre_original=a['nombre_original'],
                contenido=a['contenido'],
                tipo_mime=a['tipo_mime'],
            )

        return Response({
            'ok': True,
            'solicitud_id': sol.id,
            'archivos': [a['nombre_original'] for a in archivos_data],
            'mensaje': 'Solicitud recibida correctamente. El área de Talento Humano la procesará pronto.',
        }, status=status.HTTP_201_CREATED)


class SolicitudPSListView(APIView):
    permission_classes = [require_roles('paz_salvo', 'admin')]

    def get(self, request):
        archivos_qs = SolicitudPSArchivo.objects.only('id', 'solicitud_id', 'nombre_original', 'tipo_mime')
        solicitudes = (
            SolicitudPS.objects
            .prefetch_related(Prefetch('archivos', queryset=archivos_qs))
            .order_by('-fecha_creacion')[:200]
        )
        result = []
        for s in solicitudes:
            result.append({
                'id': s.id,
                'nombre': s.nombre,
                'identificacion': s.identificacion,
                'correo': s.correo,
                'coordinador': s.coordinador or '',
                'estado': s.estado,
                'fecha_creacion': str(s.fecha_creacion),
                'procesado_por': s.procesado_por,
                'archivos': [
                    {'id': a.id, 'nombre_original': a.nombre_original, 'tipo_mime': a.tipo_mime}
                    for a in s.archivos.all()
                ],
            })
        return Response(result)


class SolicitudPSProcesarView(APIView):
    permission_classes = [require_roles('paz_salvo', 'admin')]

    def patch(self, request, sol_id):
        try:
            sol = SolicitudPS.objects.get(id=sol_id)
        except SolicitudPS.DoesNotExist:
            return Response({'detail': 'No encontrada'}, status=status.HTTP_404_NOT_FOUND)
        sol.estado = 'PROCESADA'
        sol.procesado_por = request.pys_user['usunombre']
        sol.save(update_fields=['estado', 'procesado_por'])
        return Response({'ok': True})


class ArchivoSolicitudView(APIView):
    permission_classes = [require_roles('paz_salvo', 'admin')]

    def get(self, request, archivo_id):
        try:
            archivo = SolicitudPSArchivo.objects.get(id=archivo_id)
        except SolicitudPSArchivo.DoesNotExist:
            return Response({'detail': 'Archivo no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        contenido = bytes(archivo.contenido)
        response = HttpResponse(contenido, content_type=archivo.tipo_mime)
        response['Content-Disposition'] = f'inline; filename="{archivo.nombre_original}"'
        return response


class EncuestaRetiroExisteView(APIView):
    permission_classes = [require_roles('paz_salvo', 'admin')]

    def get(self, request, identificacion):
        enc = EncuestaRetiro.objects.filter(identificacion=identificacion).order_by('-id').first()
        if not enc:
            return Response({'existe': False})
        return Response({
            'existe': True,
            'nombre': enc.nombre,
            'fecha_retiro': str(enc.fecha_retiro) if enc.fecha_retiro else None,
            'fecha_creacion': str(enc.fecha_creacion),
        })


class EncuestaRetiroCreateView(APIView):
    """Endpoint público — no requiere autenticación."""
    permission_classes = []
    authentication_classes = []

    CALIFS_KEYS = [
        'compañeros', 'formacion', 'ambiente', 'reconocimiento',
        'carga_trabajo', 'superior', 'beneficios', 'salario',
        'valores', 'cultura', 'trabajo_equipo',
    ]
    CALIFS_VALIDOS = {'excelente', 'buena', 'mala', None, ''}

    def post(self, request):
        data = request.data
        nombre = str(data.get('nombre', '')).strip()
        identificacion = str(data.get('identificacion', '')).strip()
        correo = str(data.get('correo', '')).strip()
        fecha_retiro = data.get('fecha_retiro')

        if not nombre or not identificacion:
            return Response({'detail': 'Nombre e identificación son obligatorios'}, status=status.HTTP_400_BAD_REQUEST)
        if not identificacion.isdigit():
            return Response({'detail': 'La identificación debe contener solo números'}, status=status.HTTP_400_BAD_REQUEST)

        for key in self.CALIFS_KEYS:
            val = data.get(f'calif_{key}') or None
            if val and val not in self.CALIFS_VALIDOS:
                return Response(
                    {'detail': f'Valor inválido para calif_{key}: {val}'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        enc = EncuestaRetiro.objects.create(
            nombre=nombre,
            identificacion=identificacion,
            correo=correo,
            fecha_retiro=fecha_retiro,
            aspectos_positivos=data.get('aspectos_positivos'),
            actividades_sugeridas=data.get('actividades_sugeridas'),
            calif_compañeros=data.get('calif_compañeros') or None,
            calif_formacion=data.get('calif_formacion') or None,
            calif_ambiente=data.get('calif_ambiente') or None,
            calif_reconocimiento=data.get('calif_reconocimiento') or None,
            calif_carga_trabajo=data.get('calif_carga_trabajo') or None,
            calif_superior=data.get('calif_superior') or None,
            calif_beneficios=data.get('calif_beneficios') or None,
            calif_salario=data.get('calif_salario') or None,
            calif_valores=data.get('calif_valores') or None,
            calif_cultura=data.get('calif_cultura') or None,
            calif_trabajo_equipo=data.get('calif_trabajo_equipo') or None,
        )
        return Response({'ok': True, 'message': 'Encuesta guardada correctamente'}, status=status.HTTP_201_CREATED)


def _enc_to_dict(enc: EncuestaRetiro) -> dict:
    return {
        'id': enc.id,
        'nombre': enc.nombre,
        'identificacion': enc.identificacion,
        'correo': enc.correo,
        'fecha_retiro': str(enc.fecha_retiro) if enc.fecha_retiro else '',
        'aspectos_positivos': enc.aspectos_positivos or '',
        'actividades_sugeridas': enc.actividades_sugeridas or '',
        'calif_compañeros': enc.calif_compañeros or '',
        'calif_formacion': enc.calif_formacion or '',
        'calif_ambiente': enc.calif_ambiente or '',
        'calif_reconocimiento': enc.calif_reconocimiento or '',
        'calif_carga_trabajo': enc.calif_carga_trabajo or '',
        'calif_superior': enc.calif_superior or '',
        'calif_beneficios': enc.calif_beneficios or '',
        'calif_salario': enc.calif_salario or '',
        'calif_valores': enc.calif_valores or '',
        'calif_cultura': enc.calif_cultura or '',
        'calif_trabajo_equipo': enc.calif_trabajo_equipo or '',
        'fecha_creacion': str(enc.fecha_creacion),
    }


class EncuestaRetiroListView(APIView):
    """Lista todas las encuestas con filtro opcional por identificación."""
    permission_classes = [require_roles('paz_salvo', 'admin')]

    def get(self, request):
        qs = EncuestaRetiro.objects.order_by('-fecha_creacion')
        doc = request.query_params.get('identificacion', '').strip()
        if doc:
            qs = qs.filter(identificacion__icontains=doc)
        return Response([_enc_to_dict(e) for e in qs[:500]])


class EncuestaRetiroExportView(APIView):
    """Exporta encuestas a Excel. Sin id → todas; con ?id=X → una."""
    permission_classes = [require_roles('paz_salvo', 'admin')]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        enc_id = request.query_params.get('id', '').strip()
        doc    = request.query_params.get('identificacion', '').strip()

        qs = EncuestaRetiro.objects.order_by('-fecha_creacion')
        if enc_id:
            qs = qs.filter(id=enc_id)
        elif doc:
            qs = qs.filter(identificacion__icontains=doc)

        encuestas = list(qs[:500])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Encuestas de Retiro'

        # Estilos
        hdr_font  = Font(bold=True, color='FFFFFF', size=11)
        hdr_fill  = PatternFill('solid', fgColor='0284C7')
        hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin      = Side(style='thin', color='E2E8F0')
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt_fill  = PatternFill('solid', fgColor='F0F9FF')

        headers = [
            'ID', 'Nombre', 'Identificación', 'Correo', 'Fecha Retiro',
            'Aspectos Positivos', 'Actividades Sugeridas',
        ] + list(_ASPECT_LABELS.values()) + ['Fecha Registro']

        col_widths = [6, 30, 16, 30, 14, 40, 40] + [18] * 11 + [20]

        # Encabezados
        for col, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col, value=hdr)
            cell.font  = hdr_font
            cell.fill  = hdr_fill
            cell.alignment = hdr_align
            cell.border = cell_border
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 36

        # Datos
        for row_idx, enc in enumerate(encuestas, start=2):
            fill = alt_fill if row_idx % 2 == 0 else None
            values = [
                enc.id, enc.nombre, enc.identificacion, enc.correo,
                str(enc.fecha_retiro) if enc.fecha_retiro else '',
                enc.aspectos_positivos or '', enc.actividades_sugeridas or '',
                enc.calif_compañeros or '', enc.calif_formacion or '',
                enc.calif_ambiente or '', enc.calif_reconocimiento or '',
                enc.calif_carga_trabajo or '', enc.calif_superior or '',
                enc.calif_beneficios or '', enc.calif_salario or '',
                enc.calif_valores or '', enc.calif_cultura or '',
                enc.calif_trabajo_equipo or '',
                str(enc.fecha_creacion)[:19],
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = cell_border
                if fill:
                    cell.fill = fill
            ws.row_dimensions[row_idx].height = 20

        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f'encuestas_retiro_{enc_id or "todas"}.xlsx'
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
