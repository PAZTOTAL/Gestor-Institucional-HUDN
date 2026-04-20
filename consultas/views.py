from django.shortcuts import render, redirect
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import timedelta, datetime
from consultas_externas.models import (
    Hcnfolio, Genmedico, Hcndiapac, Gendiagno, Genpacien, Adningreso, 
    Gendetcon, Gentercer, Gentercert, Gentercerd, Genpacient, Genpaciend,
    Slnfactur, Genareser, Genespeci, Genmunici
)
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from xhtml2pdf import pisa
from django.template.loader import get_template
from core.decorators import valida_acceso
import io

def get_date_range(filter_type, start_date=None, end_date=None):
    today = timezone.localtime().date()
    if filter_type == 'diario':
        return today, today
    elif filter_type == 'semanal':
        start = today - timedelta(days=today.weekday())
        return start, today
    elif filter_type == 'mes':
        start = today.replace(day=1)
        return start, today
    elif filter_type == 'anual':
        start = today.replace(month=1, day=1)
        return start, today
    elif filter_type == 'rango' and start_date and end_date:
        return start_date, end_date
    return today, today

@valida_acceso(app_label='consultas')
def dashboard_admin(request):
    try:
        filter_type = request.GET.get('filter', 'diario')
        view_type = request.GET.get('view', 'ventas')
        group_by = request.GET.get('group_by', 'global')
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        # Debug logging
        print(f"[DEBUG] dashboard_admin called with:")
        print(f"  filter_type: {filter_type}")
        print(f"  view_type: {view_type}")
        print(f"  group_by: {group_by}")
        print(f"  start_date: {start_date_str}")
        print(f"  end_date: {end_date_str}")
        
        start_date = None
        end_date = None
        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                filter_type = 'rango'
            except ValueError as e:
                print(f"[ERROR] Date parsing error: {e}")
                pass

        sd, ed = get_date_range(filter_type, start_date, end_date)
        print(f"[DEBUG] Date range: {sd} to {ed}")
        
        # Base query for invoices
        try:
            facturas_query = Slnfactur.objects.using('readonly').filter(
                sfafecfac__date__range=[sd, ed],
                sfadocanu=False
            )
            
            total_facturado = facturas_query.aggregate(total=Sum('sfatotfac'))['total'] or 0
            conteo_facturas = facturas_query.count()
            rips_count = facturas_query.count()
            
            print(f"[DEBUG] Query results: {conteo_facturas} facturas, total: ${total_facturado}")
            
        except Exception as e:
            print(f"[ERROR] Database query error: {e}")
            # Return safe defaults
            total_facturado = 0
            conteo_facturas = 0
            rips_count = 0
            facturas_query = Slnfactur.objects.using('readonly').none()

        facturas_detalle = []
        
        try:
            if group_by == 'aseguradora':
                print(f"[DEBUG] Grouping by aseguradora")
                facturas_detalle = facturas_query.values(
                    'gendetcon__gdecodigo', 'gendetcon__gdenombre'
                ).annotate(
                    total=Sum('sfatotfac'),
                    cantidad=Count('oid')
                ).order_by('-total')
            elif group_by == 'paciente':
                print(f"[DEBUG] Grouping by paciente")
                facturas_detalle = facturas_query.values(
                    'adningreso__genpacien__pacnumdoc', 
                    'adningreso__genpacien__pactipdoc', 
                    'adningreso__genpacien__pacprinom', 
                    'adningreso__genpacien__pacpriape'
                ).annotate(
                    total=Sum('sfatotfac'),
                    cantidad=Count('oid')
                ).order_by('-total')[:200]
            elif group_by == 'factura':
                print(f"[DEBUG] Grouping by factura")
                facturas_detalle = facturas_query.select_related('adningreso__genpacien', 'gendetcon').order_by('-sfafecfac')[:200]
            else:
                print(f"[DEBUG] Using global grouping (no detail)")
                facturas_detalle = []
                
            print(f"[DEBUG] Retrieved {len(list(facturas_detalle)) if facturas_detalle else 0} detail records")
            
        except Exception as e:
            print(f"[ERROR] Grouping query error: {e}")
            import traceback
            traceback.print_exc()
            facturas_detalle = []

        # Insurers for selectivity if needed in the future
        try:
            aseguradoras = Gendetcon.objects.using('readonly').values('oid', 'gdenombre', 'gdecodigo').order_by('gdenombre')[:100]
        except Exception as e:
            print(f"[ERROR] Error fetching aseguradoras: {e}")
            aseguradoras = []

        context = {
            'filter_type': filter_type,
            'view_type': view_type,
            'group_by': group_by,
            'sd': sd,
            'ed': ed,
            'total_facturado': total_facturado,
            'conteo_facturas': conteo_facturas,
            'rips_count': rips_count,
            'aseguradoras': aseguradoras,
            'facturas_detalle': facturas_detalle,
        }
        
        print(f"[DEBUG] Rendering template with context")
        return render(request, 'consultas/dashboard_admin.html', context)
        
    except Exception as e:
        print(f"[CRITICAL ERROR] Unhandled exception in dashboard_admin: {e}")
        import traceback
        traceback.print_exc()
        # Return error page or redirect
        from django.http import HttpResponse
        return HttpResponse(f"Error en dashboard_admin: {str(e)}", status=500)

@valida_acceso(app_label='consultas')
def dashboard_salud(request):
    filter_type = request.GET.get('filter', 'diario')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    # Filtros adicionales
    aseguradora_id = request.GET.get('aseguradora')
    area_id = request.GET.get('area')
    municipio_id = request.GET.get('municipio')
    sexo = request.GET.get('sexo')
    especialidad_id = request.GET.get('especialidad')

    start_date = None
    end_date = None
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        filter_type = 'rango'

    sd, ed = get_date_range(filter_type, start_date, end_date)

    # Pacientes atendidos (basado en facturas o folios de historia clínica)
    # Usaremos Slnfactur como base para "atenciones facturadas"
    query = Q(sfafecfac__date__range=[sd, ed], sfadocanu=False)
    
    if aseguradora_id:
        query &= Q(gendetcon=aseguradora_id)
    # Nota: para filtrar por Area, Sexo, etc., necesitamos joins o subconsultas si no están en Slnfactur
    # Por ahora, implementaremos el conteo básico y los filtros simples.
    
    atenciones = Slnfactur.objects.using('readonly').filter(query)
    total_pacientes = atenciones.count()

    # Listas para los selectores de filtros
    aseguradoras = Gendetcon.objects.using('readonly').values('oid', 'gdenombre').order_by('gdenombre')[:100]
    areas = Genareser.objects.using('readonly').values('oid', 'gasnombre').order_by('gasnombre')
    especialidades = Genespeci.objects.using('readonly').values('oid', 'geedescri').order_by('geedescri')
    municipios = Genmunici.objects.using('readonly').values('oid', 'munnommun').order_by('munnommun')[:100]

    context = {
        'filter_type': filter_type,
        'sd': sd,
        'ed': ed,
        'total_pacientes': total_pacientes,
        'aseguradoras': aseguradoras,
        'areas': areas,
        'especialidades': especialidades,
        'municipios': municipios,
    }
    return render(request, 'consultas/dashboard_salud.html', context)

@valida_acceso(app_label='consultas')
def dashboard_aseguradoras(request):
    today = timezone.localtime().date()
    # Ranges
    in_day = today
    in_week = today - timedelta(days=today.weekday())
    in_month = today.replace(day=1)
    in_year = today.replace(month=1, day=1)

    # We fetch all active insurers
    aseguradoras = Gendetcon.objects.using('readonly').all().order_by('gdenombre')[:150]
    
    # We'll calculate stats for each (can be optimized but this is readable)
    # Using filter by the integer ID gendetcon in Slnfactur
    stats = []
    for ase in aseguradoras:
        base_query = Slnfactur.objects.using('readonly').filter(gendetcon=ase.oid, sfadocanu=False)
        
        dia = base_query.filter(sfafecfac__date=in_day).count()
        sem = base_query.filter(sfafecfac__date__range=[in_week, today]).count()
        mes = base_query.filter(sfafecfac__date__range=[in_month, today]).count()
        ano = base_query.filter(sfafecfac__date__range=[in_year, today]).count()
        
        if any([dia, sem, mes, ano]):
            stats.append({
                'nombre': ase.gdenombre,
                'codigo': ase.gdecodigo,
                'dia': dia,
                'sem': sem,
                'mes': mes,
                'ano': ano,
            })
    
    # Sort by month by default
    stats = sorted(stats, key=lambda x: x['mes'], reverse=True)

    context = {
        'stats': stats,
        'today': today,
        'aseguradoras_list': Gendetcon.objects.using('readonly').values('oid', 'gdenombre', 'gdecodigo').order_by('gdenombre')[:150],
    }
    return render(request, 'consultas/dashboard_aseguradoras.html', context)

@valida_acceso(app_label='consultas')
def exportar_pdf(request):
    # Lógica para exportar el dashboard actual a PDF
    # Se basará en el tipo de reporte solicitado
    reporte = request.GET.get('tipo', 'admin')
    
    # Re-ejecutar lógica de consulta o pasar datos (simplificado: re-ejecutar)
    if reporte == 'admin':
        # ... lógica similar a dashboard_admin ...
        template_path = 'consultas/pdf_admin.html'
        context = {} # Poblar con datos reales
    else:
        template_path = 'consultas/pdf_salud.html'
        context = {} # Poblar con datos reales

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_{reporte}.pdf"'
    
    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
       return HttpResponse('Error al generar PDF', status=500)
    return response


from .utils import RipsGenerator
from datetime import datetime, timedelta

@valida_acceso(app_label='consultas')
def informes_rips(request):
    return render(request, 'consultas/informes_rips.html')

@valida_acceso(app_label='consultas')
def generar_rips(request):
    if request.method == 'POST':
        fecha_ini = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        
        # Parse dates strings 'YYYY-MM-DD'
        try:
            start_date = datetime.strptime(fecha_ini, '%Y-%m-%d')
            end_date = datetime.strptime(fecha_fin, '%Y-%m-%d')
            
            # For end date, maybe set to end of day? 
            # But filter usually works with dates if field is DateTimeField might need adjustment.
            # Assuming 'sfafecfac' is DateTime, standard filter might need range to cover full day.
            # Let's adjust end_date to end of day
            end_date = end_date.replace(hour=23, minute=59, second=59)
            
            generator = RipsGenerator(start_date, end_date)
            zip_buffer = generator.generate()
            
            if not zip_buffer:
                # Handle no data case - maybe flash message or return empty
                # For now just redirect with error param
                return redirect('consultas:dashboard_admin') 
                
            response = HttpResponse(zip_buffer, content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename=RIPS_{fecha_ini}_{fecha_fin}.zip'
            return response
            
        except ValueError:
            pass # Handle invalid dates

    return redirect('consultas:dashboard_admin')




@valida_acceso(app_label='consultas')
def produccion_medico(request):
    fecha_inicio_str = request.GET.get('fecha_inicio', datetime.now().strftime('%Y-%m-%d'))
    fecha_fin_str = request.GET.get('fecha_fin', datetime.now().strftime('%Y-%m-%d'))
    doctor_ids = request.GET.getlist('doctor_ids')
    discrete_dates_str = request.GET.get('fechas_especificas', '')
    filtrar = request.GET.get('filtrar')

    atenciones_por_medico = {}
    doctors_list = Genmedico.objects.using('readonly').all().order_by('gmenomcom')
    
    if filtrar:
        try:
            start_date = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            end_date = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)

            # Specific dates filter
            discrete_dates = []
            if discrete_dates_str:
                import re
                parts = re.split(r'[,;\s]+', discrete_dates_str)
                for p in parts:
                    if not p: continue
                    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                        try:
                            d_obj = datetime.strptime(p, fmt).date()
                            discrete_dates.append(d_obj)
                            break
                        except ValueError: continue

            if discrete_dates:
                query = Hcnfolio.objects.using('readonly').filter(hcfecfol__date__in=discrete_dates)
            else:
                query = Hcnfolio.objects.using('readonly').filter(hcfecfol__range=[start_date, end_date])
            
            # Doctor filtering fix
            clean_doctor_ids = [int(d_id) for d_id in doctor_ids if d_id.isdigit()]
            if clean_doctor_ids:
                query = query.filter(genmedico__in=clean_doctor_ids)

            folios = query.order_by('hcfecfol') 

            paciente_ids = set()
            medico_ids = set()
            folio_ids = []

            for f in folios:
                if f.genpacien_id:
                    paciente_ids.add(f.genpacien_id)
                mid = getattr(f, 'genmedico_id', None)
                if mid:
                   medico_ids.add(mid)
                folio_ids.append(f.oid)

            pacientes = {p.oid: p for p in Genpacien.objects.using('readonly').filter(oid__in=paciente_ids)}
            medicos = {m.oid: m for m in Genmedico.objects.using('readonly').filter(oid__in=medico_ids)}
            
            adningreso_ids = set()
            for f in folios:
                if f.adningreso:
                    adningreso_ids.add(f.adningreso)
            
            adningresos = {a.oid: a for a in Adningreso.objects.using('readonly').filter(oid__in=adningreso_ids)}
            
            start_search = start_date - timedelta(days=30)
            extra_admissions = Adningreso.objects.using('readonly').filter(
                genpacien__in=paciente_ids, 
                ainfecing__lte=end_date,
                ainfecing__gte=start_search
            ).order_by('ainfecing')
            
            admissions_by_patient = {}
            diag_ids = set()
            for adm in extra_admissions:
                if adm.genpacien_id not in admissions_by_patient:
                    admissions_by_patient[adm.genpacien_id] = []
                admissions_by_patient[adm.genpacien_id].append(adm)
                if adm.dgndiagno:
                    diag_ids.add(adm.dgndiagno)

            diags_pac = Hcndiapac.objects.using('readonly').filter(hcnfolio__in=folio_ids)
            folio_diag_map = {}

            for dp in diags_pac:
                if dp.gendiagno:
                    diag_ids.add(dp.gendiagno)
                    
            for a in adningresos.values():
                if a.dgndiagno:
                    diag_ids.add(a.dgndiagno)

            diagnosticos_master = {d.oid: d for d in Gendiagno.objects.using('readonly').filter(oid__in=diag_ids)}

            for dp in diags_pac:
                 if dp.gendiagno in diagnosticos_master:
                     folio_diag_map[dp.hcnfolio_id] = diagnosticos_master[dp.gendiagno]

            gendetcon_ids = set()
            for adm in adningresos.values():
                if adm.gendetcon_id:
                     gendetcon_ids.add(adm.gendetcon_id)
            for adm_list in admissions_by_patient.values():
                for adm in adm_list:
                    if adm.gendetcon_id:
                        gendetcon_ids.add(adm.gendetcon_id)
                     

            target_planes = {p.oid: p for p in Gendetcon.objects.using('readonly').filter(oid__in=gendetcon_ids)}
            
            terceros_ids = set()
            for adm in adningresos.values():
                if adm.entidadadministradora:
                    terceros_ids.add(adm.entidadadministradora)
            for adm_list in admissions_by_patient.values():
                for adm in adm_list:
                    if adm.entidadadministradora:
                        terceros_ids.add(adm.entidadadministradora)
            
            terceros = {t.oid: t for t in Gentercer.objects.using('readonly').filter(oid__in=terceros_ids)}

            # Patient contact details fallbacks
            patient_tercer_ids = [p.gentercer for p in pacientes.values() if p.gentercer]
            
            # Fetch from Genpacient/d (direct patient details)
            p_tels = {t.genpacien: t.pactelefono for t in Genpacient.objects.using('readonly').filter(genpacien__in=paciente_ids).order_by('-pactelprinc')}
            p_dirs = {d.genpacien: d.pacdireccion for d in Genpaciend.objects.using('readonly').filter(genpacien__in=paciente_ids).order_by('-pacdiprinc')}
            
            # Fetch from Gentercert/d (terceros details)
            t_tels = {t.gentercer: t.tertelefono for t in Gentercert.objects.using('readonly').filter(gentercer__in=patient_tercer_ids).order_by('-tel_princi' if 'tel_princi' in [f.name for f in Gentercert._meta.fields] else '-telprinci')}
            t_dirs = {d.gentercer: d.terdireccion for d in Gentercerd.objects.using('readonly').filter(gentercer__in=patient_tercer_ids).order_by('-dir_princi' if 'dir_princi' in [f.name for f in Gentercerd._meta.fields] else '-dirprinci')}



            # Gender mapping
            sex_map = {1: 'M', 2: 'F'}

            for folio in folios:
                medico_id = getattr(folio, 'genmedico', None)
                if not medico_id or medico_id not in medicos:
                     continue
                
                medico = medicos[medico_id]
                paciente = pacientes.get(folio.genpacien)
            
                diag_obj = folio_diag_map.get(folio.oid)
                current_adm = None
                
                if not diag_obj and folio.adningreso:
                    current_adm = adningresos.get(folio.adningreso)
                    if current_adm and current_adm.dgndiagno:
                        diag_obj = diagnosticos_master.get(current_adm.dgndiagno)
                
                if not current_adm and folio.genpacien in admissions_by_patient:
                    folio_date = folio.hcfecfol
                    candidates = admissions_by_patient[folio.genpacien]
                    for adm in reversed(candidates):
                        try:
                            f_date = folio_date.replace(tzinfo=None) if folio_date else None
                            a_date = adm.ainfecing.replace(tzinfo=None) if adm.ainfecing else None
                            if f_date and a_date:
                                delta = a_date - f_date
                                is_valid_start = (a_date <= f_date) or (delta.total_seconds() < 86400 and delta.total_seconds() >= 0)
                                if is_valid_start:
                                    current_adm = adm 
                                    if not diag_obj and adm.dgndiagno:
                                        diag_obj = diagnosticos_master.get(adm.dgndiagno)
                                    break
                        except: pass
                
                edad_str = ""
                if paciente and paciente.gpafecnac:
                    try:
                        born = paciente.gpafecnac
                        if hasattr(born, 'date'): born = born.date()
                        today = folio.hcfecfol.date()
                        age_years = today.year - born.year - ((today.month, today.day) < (born.month, born.day))
                        if age_years > 0:
                            edad_str = f"{age_years} Años"
                        else:
                             delta = today - born
                             if delta.days > 30:
                                 edad_str = f"{int(delta.days/30)} Meses"
                             else:
                                 edad_str = f"{delta.days} Días"
                    except: pass
                
                eps_nombre = ""
                asegurador_nombre = ""
                if current_adm:
                    if current_adm.gendetcon_id in target_planes:
                         eps_nombre = target_planes[current_adm.gendetcon_id].gdenombre
                    if current_adm.entidadadministradora in terceros:
                         asegurador_nombre = f"{terceros[current_adm.entidadadministradora].terprinom or ''} {terceros[current_adm.entidadadministradora].terpriape or ''}".strip()
                         # Fallback for search/display if only one name field exists or something
                         if not asegurador_nombre and hasattr(terceros[current_adm.entidadadministradora], 'ternomcom'):
                             asegurador_nombre = terceros[current_adm.entidadadministradora].ternomcom

                if medico not in atenciones_por_medico:
                    atenciones_por_medico[medico] = []
                
                celular = paciente.gpatelresex if (paciente and paciente.gpatelresex and paciente.gpatelresex != "-") else p_tels.get(folio.genpacien)
                if not celular or celular == "-": 
                    celular = t_tels.get(paciente.gentercer) if paciente else "-"
                
                direccion = paciente.gpadirresex if (paciente and paciente.gpadirresex and paciente.gpadirresex != "-") else p_dirs.get(folio.genpacien)
                if not direccion or direccion == "-":
                    direccion = t_dirs.get(paciente.gentercer) if paciente else "-"
                if (not direccion or direccion == "-") and paciente:
                    direccion = getattr(paciente, 'gpadirrhab', "-")

                atenciones_por_medico[medico].append({
                    'folio': folio,
                    'paciente': paciente, 
                    'edad': edad_str,
                    'eps': eps_nombre,
                    'ingreso_id': current_adm.oid if current_adm else (folio.adningreso or "No Link"),
                    'ingreso_fecha': current_adm.ainfecing.strftime('%Y-%m-%d %H:%M') if current_adm and current_adm.ainfecing else "-",
                    'diagnostico': diag_obj,
                    'diagnostico_nombre': diag_obj.dianombre if diag_obj else "Sin diagnóstico",
                    'diagnostico_codigo': diag_obj.diacodigo if diag_obj else "-",
                    'asegurador': asegurador_nombre,
                    'sexo': sex_map.get(paciente.gpasexpac if paciente else None, "-"),
                    'celular': celular or "-",
                    'direccion': direccion or "-",
                })

        except ValueError:
            pass

    context = {
        'atenciones_por_medico': atenciones_por_medico,
        'fecha_inicio_str': fecha_inicio_str,
        'fecha_fin_str': fecha_fin_str,
        'doctors_list': doctors_list,
        'selected_doctors': [int(d_id) for d_id in doctor_ids if d_id.isdigit()],
        'fechas_especificas': discrete_dates_str,
        'is_filtered': bool(filtrar)
    }
    return render(request, 'consultas/produccion_medico.html', context)

@valida_acceso(app_label='consultas')
def produccion_medico_excel(request):
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    doctor_ids = request.GET.getlist('doctor_ids')
    discrete_dates_str = request.GET.get('fechas_especificas', '')

    if not fecha_inicio_str or not fecha_fin_str:
        return HttpResponse("Faltan parámetros de fecha", status=400)

    try:
        start_date = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
        end_date = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
        end_date = end_date.replace(hour=23, minute=59, second=59)

        discrete_dates = []
        if discrete_dates_str:
            import re
            parts = re.split(r'[,;\s]+', discrete_dates_str)
            for p in parts:
                if not p: continue
                for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                    try:
                        d_obj = datetime.strptime(p, fmt).date()
                        discrete_dates.append(d_obj)
                        break
                    except ValueError: continue

        if discrete_dates:
            query = Hcnfolio.objects.filter(hcfecfol__date__in=discrete_dates)
        else:
            query = Hcnfolio.objects.filter(hcfecfol__range=[start_date, end_date])

        clean_doctor_ids = [int(d_id) for d_id in doctor_ids if d_id.isdigit()]
        if clean_doctor_ids:
            query = query.filter(genmedico__in=clean_doctor_ids)

        folios = query.order_by('hcfecfol')

        # Reuse logic from produccion_medico (simplified for Excel)
        paciente_ids = set()
        medico_ids = set()
        folio_ids = []
        for f in folios:
            if f.genpacien: paciente_ids.add(f.genpacien)
            mid = getattr(f, 'genmedico', None)
            if mid: medico_ids.add(mid)
            folio_ids.append(f.oid)

        pacientes = {p.oid: p for p in Genpacien.objects.filter(oid__in=paciente_ids)}
        medicos = {m.oid: m for m in Genmedico.objects.filter(oid__in=medico_ids)}
        
        adningreso_ids = set([f.adningreso for f in folios if f.adningreso])
        adningresos = {a.oid: a for a in Adningreso.objects.filter(oid__in=adningreso_ids)}
        
        # Diag lookup
        diags_pac = Hcndiapac.objects.filter(hcnfolio__in=folio_ids)
        diag_ids = set([dp.gendiagno for dp in diags_pac if dp.gendiagno])
        for a in adningresos.values():
            if a.dgndiagno: diag_ids.add(a.dgndiagno)
        diagnosticos_master = {d.oid: d for d in Gendiagno.objects.filter(oid__in=diag_ids)}
        
        folio_diag_map = {}
        for dp in diags_pac:
             if dp.gendiagno in diagnosticos_master:
                 folio_diag_map[dp.hcnfolio_id] = diagnosticos_master[dp.gendiagno]

        gendetcon_ids = set([a.gendetcon_id for a in adningresos.values() if a.gendetcon_id])
        target_planes = {p.oid: p for p in Gendetcon.objects.filter(oid__in=gendetcon_ids)}
        
        terceros_ids = set([a.entidadadministradora for a in adningresos.values() if a.entidadadministradora])
        terceros = {t.oid: t for t in Gentercer.objects.filter(oid__in=terceros_ids)}

        # Patient contact details fallbacks for Excel
        patient_tercer_ids = [p.gentercer for p in pacientes.values() if p.gentercer]
        p_tels = {t.genpacien: t.pactelefono for t in Genpacient.objects.filter(genpacien__in=paciente_ids).order_by('-pactelprinc')}
        p_dirs = {d.genpacien: d.pacdireccion for d in Genpaciend.objects.filter(genpacien__in=paciente_ids).order_by('-pacdiprinc')}
        t_tels = {t.gentercer: t.tertelefono for t in Gentercert.objects.filter(gentercer__in=patient_tercer_ids).order_by('-telprinci')}
        t_dirs = {d.gentercer: d.terdireccion for d in Gentercerd.objects.filter(gentercer__in=patient_tercer_ids).order_by('-dirprinci')}

        # Workbook creation
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Producción Médica"

        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        headers = [
            "N°", "NOMBRE", "CEDULA", "FECHA NACIMIENTO", "EDAD", 
            "CELULAR", "DIRECCION", "FECHA ATENCIÓN", "SEXO", 
            "DIAGNOSTICO", "ASEGURADOR", "PLAN", "MEDICO ESPEC"
        ]

        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        sex_map = {1: 'M', 2: 'F'}
        row_idx = 2
        for i, folio in enumerate(folios, 1):
            paciente = pacientes.get(folio.genpacien)
            medico = medicos.get(getattr(folio, 'genmedico', None))
            adm = adningresos.get(folio.adningreso)
            plan = target_planes.get(adm.gendetcon_id).gdenombre if adm and adm.gendetcon_id in target_planes else "-"
            eps = "-"
            if adm and adm.entidadadministradora in terceros:
                t = terceros[adm.entidadadministradora]
                eps = f"{t.terprinom or ''} {t.terpriape or ''}".strip() or getattr(t, 'ternomcom', "-")
            
            diag = folio_diag_map.get(folio.oid) or (diagnosticos_master.get(adm.dgndiagno) if adm else None)
            
            # Age calc
            edad_str = ""
            if paciente and paciente.gpafecnac:
                try:
                    born = paciente.gpafecnac.date() if hasattr(paciente.gpafecnac, 'date') else paciente.gpafecnac
                    today = folio.hcfecfol.date()
                    age = today.year - born.year - ((today.month, today.day) < (born.month, born.day))
                    edad_str = f"{age} Años" if age > 0 else "Menor a 1 Año"
                except: pass

            # Contact resolution
            celular = paciente.gpatelresex if (paciente and paciente.gpatelresex and paciente.gpatelresex != "-") else p_tels.get(folio.genpacien)
            if not celular or celular == "-": celular = t_tels.get(paciente.gentercer) if paciente else "-"
            
            direccion = paciente.gpadirresex if (paciente and paciente.gpadirresex and paciente.gpadirresex != "-") else p_dirs.get(folio.genpacien)
            if not direccion or direccion == "-": direccion = t_dirs.get(paciente.gentercer) if paciente else "-"
            if (not direccion or direccion == "-") and paciente: direccion = getattr(paciente, 'gpadirrhab', "-")

            ws.cell(row=row_idx, column=1, value=i).border = border
            ws.cell(row=row_idx, column=2, value=f"{paciente.pacprinom} {paciente.pacsegnom or ''} {paciente.pacpriape} {paciente.pacsegape or ''}".strip()).border = border
            ws.cell(row=row_idx, column=3, value=paciente.pacnumdoc if paciente else "-").border = border
            ws.cell(row=row_idx, column=4, value=paciente.gpafecnac.strftime('%d/%m/%Y') if paciente and paciente.gpafecnac else "-").border = border
            ws.cell(row=row_idx, column=5, value=edad_str).border = border
            ws.cell(row=row_idx, column=6, value=celular or "-").border = border
            ws.cell(row=row_idx, column=7, value=direccion or "-").border = border
            ws.cell(row=row_idx, column=8, value=folio.hcfecfol.strftime('%d/%m/%Y %H:%M')).border = border
            ws.cell(row=row_idx, column=9, value=sex_map.get(paciente.gpasexpac if paciente else None, "-")).border = border
            ws.cell(row=row_idx, column=10, value=diag.diacodigo if diag else "Sin diagnóstico").border = border
            ws.cell(row=row_idx, column=11, value=eps).border = border
            ws.cell(row=row_idx, column=12, value=plan).border = border
            ws.cell(row=row_idx, column=13, value=medico.gmenomcom if medico else "-").border = border
            row_idx += 1

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = max_length + 2

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Producción_Médica_{fecha_inicio_str}_{fecha_fin_str}.xlsx'
        wb.save(response)
        return response

    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)



# ============================================
# PATIENT TRACKING VIEWS
# ============================================

from consultas_externas.models import Adningreso, Hcnfolio, Hcndiapac, Gendiagno, Hcnnotenf, Hcnregenf
from django.shortcuts import get_object_or_404


@valida_acceso(app_label='consultas')
def pacientes_urgencias_list(request):
    """
    Lista de pacientes activos en TODO el hospital (no solo Urgencias)
    Solo muestra resultados cuando el usuario busca
    """
    # Buscar parámetro de búsqueda
    search_query = request.GET.get('search', None)
    
    # Inicializar lista vacía
    pacientes_data = []
    pacientes_por_area = {}
    total_pacientes = 0
    searched = False
    
    # Si 'search' está en los parámetros GET, significa que el usuario hizo clic en Buscar
    if search_query is not None:
        searched = True
        search_query = search_query.strip()
        
        # Query base: pacientes activos en TODO el hospital
        query = Q(ainestado=1)
        
        # Solo filtrar si hay un término de búsqueda, si no, mostrar todos los activos
        if search_query:
            # Buscar por documento o nombre del paciente
            query &= (
                Q(genpacien__pacnumdoc__icontains=search_query) |
                Q(genpacien__pacprinom__icontains=search_query) |
                Q(genpacien__pacsegnom__icontains=search_query) |
                Q(genpacien__pacpriape__icontains=search_query) |
                Q(genpacien__pacsegape__icontains=search_query)
            )
        
        # Obtener ingresos activos con datos relacionados
        ingresos = Adningreso.objects.using('readonly').filter(query).select_related(
            'genpacien',
            'gendetcon'
        ).order_by('-ainfecing')[:100]  # Limitar a 100 resultados cuando se busca
        
        # Enriquecer con información adicional y agrupar por área
        pacientes_por_area = {}
        for ingreso in ingresos:
            # Obtener el folio más reciente para este ingreso
            ultimo_folio = Hcnfolio.objects.using('readonly').filter(
                adningreso=ingreso.oid
            ).order_by('-hcfecfol').first()
            
            # Contar diagnósticos
            num_diagnosticos = 0
            ubicacion_actual = 'Sin ubicación registrada'
            
            if ultimo_folio:
                num_diagnosticos = Hcndiapac.objects.using('readonly').filter(
                    hcnfolio=ultimo_folio.oid
                ).count()
                
                # Obtener ubicación actual (área del último folio)
                if hasattr(ultimo_folio, 'genareser') and ultimo_folio.genareser:
                    if hasattr(ultimo_folio.genareser, 'gasnombre'):
                        ubicacion_actual = ultimo_folio.genareser.gasnombre
                    else:
                        ubicacion_actual = f'Área ID: {ultimo_folio.genareser}'
            
            p_data = {
                'ingreso': ingreso,
                'paciente': ingreso.genpacien if hasattr(ingreso, 'genpacien') else None,
                'ultimo_folio': ultimo_folio,
                'num_diagnosticos': num_diagnosticos,
                'dias_ingresado': (timezone.now() - ingreso.ainfecing).days if ingreso.ainfecing else 0,
                'ubicacion_actual': ubicacion_actual,
            }
            
            if ubicacion_actual not in pacientes_por_area:
                pacientes_por_area[ubicacion_actual] = []
            pacientes_por_area[ubicacion_actual].append(p_data)
            pacientes_data.append(p_data)
        
        total_pacientes = len(pacientes_data)
        # Ordenar áreas alfabéticamente
        pacientes_por_area = dict(sorted(pacientes_por_area.items()))
    
    context = {
        'pacientes_por_area': pacientes_por_area,
        'search_query': search_query if search_query is not None else '',
        'total_pacientes': total_pacientes,
        'searched': searched,
    }
    
    return render(request, 'consultas/pacientes_urgencias_list.html', context)


@valida_acceso(app_label='consultas')
def paciente_detalle(request, ingreso_id):
    """
    Vista detallada del recorrido completo de un paciente
    """
    # Obtener el ingreso
    ingreso = get_object_or_404(
        Adningreso.objects.select_related('genpacien', 'gendetcon'),
        oid=ingreso_id
    )
    
    paciente = ingreso.genpacien if hasattr(ingreso, 'genpacien') else None
    
    # Obtener todos los folios (evoluciones) de este ingreso
    folios = Hcnfolio.objects.filter(
        adningreso=ingreso_id
    ).order_by('hcfecfol')
    
    # Construir timeline de ubicaciones y eventos
    timeline = []
    diagnosticos_set = set()
    diagnosticos_list = []
    
    for folio in folios:
        # Obtener diagnósticos de este folio
        diags_folio = Hcndiapac.objects.filter(
            hcnfolio=folio.oid
        )
        
        folio_diagnosticos = []
        try:
            for diag_pac in diags_folio:
                # Obtener detalle del diagnóstico (Gendiagno)
                diag_detalle = None
                if diag_pac.gendiagno:
                    diag_detalle = Gendiagno.objects.using('readonly').filter(oid=diag_pac.gendiagno).first()
                    
                if diag_detalle:
                    nombre_diag = f"{diag_detalle.dianocodigo} - {diag_detalle.dianonombre}" if diag_detalle.dianocodigo else diag_detalle.dianonombre
                else:
                    nombre_diag = "Diagnóstico ID: " + str(diag_pac.gendiagno)

                if diag_pac.gendiagno and diag_pac.gendiagno not in diagnosticos_set:
                    diagnosticos_set.add(diag_pac.gendiagno)
                    diagnosticos_list.append({
                        'diagnostico': diag_pac.gendiagno,
                        'fecha': folio.hcfecfol,
                        'tipo': 'Principal' if diag_pac.hcpdiaprin else 'Secundario',
                        'nombre': nombre_diag
                    })
                
                folio_diagnosticos.append({
                    'nombre': nombre_diag,
                    'tipo': 'Principal' if diag_pac.hcpdiaprin else 'Secundario'
                })
        except Exception as e:
            print(f"Error fetching diagnoses: {e}")
        
        # Agregar evento al timeline
        # genareser puede ser un ID (int) o un objeto relacionado
        area_nombre = 'Área no especificada'
        if hasattr(folio, 'genareser') and folio.genareser:
            if hasattr(folio.genareser, 'gasnombre'):
                area_nombre = folio.genareser.gasnombre
            else:
                # Es un ID, mostrar solo el ID
                area_nombre = f'Área ID: {folio.genareser}'
        
        timeline.append({
            'fecha': folio.hcfecfol,
            'tipo': 'evolucion',
            'descripcion': f'Evolución en {area_nombre}',
            'area': area_nombre,
            'folio_id': folio.oid,
            'diagnosticos': folio_diagnosticos
        })
    
    # --- FETCH NURSING NOTES (Hcnnotenf) ---
    # Logic: Search for Hcnregenf records for this patient within the admission time range
    # Then find Hcnnotenf records linked to those Hcnregenf IDs.
    if paciente:
        try:
            # Define time window: Admission Start to Now (or Discharge Date if exists)
            start_date = ingreso.ainfecing
            end_date = ingreso.ainfecegre if ingreso.ainfecegre else timezone.now()
            
            # Get Nursing Registers for this patient in range
            nursing_regs = Hcnregenf.objects.using('readonly').filter(
                genpacien=paciente.oid,
                hcfecreg__gte=start_date,
                hcfecreg__lte=end_date
            ).values_list('oid', flat=True)
            
            if nursing_regs:
                # Get the actual notes
                notas_enfermeria = Hcnnotenf.objects.using('readonly').filter(
                    hcnregenf__in=list(nursing_regs)
                ) # Removed select_related as it might not be needed or supported if relations aren't clear
                
                for nota in notas_enfermeria:
                    # Find timestamp from Hcnregenf
                    try:
                        # Optimization: Fetch regenf with date
                        reg = Hcnregenf.objects.using('readonly').get(oid=nota.hcnregenf)
                        
                        timeline.append({
                            'fecha': reg.hcfecreg,
                            'tipo': 'nota_enfermeria',
                            'descripcion': nota.hcntitulo or 'Nota de Enfermería',
                            'contenido': nota.hcndescrip or '',
                            'area': 'Enfermería', 
                        })
                    except Exception as e_inner:
                        print(f"Error checking nursing note reg: {e_inner}")
                        continue
        except Exception as e:
            print(f"Error fetching nursing notes: {e}")

    # Agregar evento de ingreso al inicio del timeline
    timeline.insert(0, {
        'fecha': ingreso.ainfecing,
        'tipo': 'ingreso',
        'descripcion': 'Ingreso a Urgencias',
        'area': 'Triage / Urgencias',
    })
    
    # Ordenar timeline por fecha
    timeline.sort(key=lambda x: x['fecha'] if x['fecha'] else timezone.now())
    
    # Calcular estadísticas
    dias_hospitalizacion = (timezone.now() - ingreso.ainfecing).days if ingreso.ainfecing else 0
    horas_hospitalizacion = (timezone.now() - ingreso.ainfecing).total_seconds() / 3600 if ingreso.ainfecing else 0
    
    # Calcular edad del paciente
    edad_str = ""
    if paciente and paciente.gpafecnac:
        try:
            born = paciente.gpafecnac
            if hasattr(born, 'date'):
                born = born.date()
            today = timezone.now().date()
            
            age_years = today.year - born.year - ((today.month, today.day) < (born.month, born.day))
            if age_years > 0:
                edad_str = f"{age_years} años"
            else:
                delta = today - born
                if delta.days > 30:
                    edad_str = f"{int(delta.days/30)} meses"
                else:
                    edad_str = f"{delta.days} días"
        except:
            pass
    
    context = {
        'ingreso': ingreso,
        'paciente': paciente,
        'edad': edad_str,
        'timeline': timeline,
        'diagnosticos': diagnosticos_list,
        'num_evoluciones': folios.count(),
        'dias_hospitalizacion': dias_hospitalizacion,
        'horas_hospitalizacion': round(horas_hospitalizacion, 1),
        'estado_actual': 'Activo' if ingreso.ainestado == 1 else 'Cerrado',
    }
    
    return render(request, 'consultas/paciente_detalle.html', context)


@valida_acceso(app_label='consultas')
def paciente_historial(request, documento=None):
    """
    Vista de historial clínico vitalicio e indicadores por paciente
    """
    if not documento:
        documento = request.GET.get('documento')
    
    if not documento:
        return render(request, 'consultas/paciente_historial.html', {'searched': False})
        
    # 1. Obtener Paciente
    paciente = Genpacien.objects.using('readonly').filter(pacnumdoc__icontains=documento).first()
    if not paciente:
        paciente = Genpacien.objects.using('readonly').filter(pacnumdoc=documento).first()
        
    if not paciente:
        return render(request, 'consultas/paciente_historial.html', {
            'searched': True, 
            'found': False, 
            'documento': documento
        })

    # 2. Obtener todos sus ingresos (Historial Vitalicio)
    ingresos_qs = Adningreso.objects.using('readonly').filter(genpacien=paciente.oid).order_by('-ainfecing')
    
    # 3. Procesar Ingresos e Indicadores
    total_ingresos = ingresos_qs.count()
    tiempos_atencion = [] 
    total_dias_estancia = 0
    historial_data = []
    conteo_diagnosticos = {}
    
    for ing in ingresos_qs:
        # Buscar folios para este ingreso
        folios = Hcnfolio.objects.using('readonly').filter(adningreso=ing.oid).order_by('hcfecfol')
        primer_folio = folios.first()
        
        # Oportunidad de Atención (Minutos)
        oportunidad = None
        if ing.ainfecing and primer_folio and primer_folio.hcfecfol:
            diff = primer_folio.hcfecfol - ing.ainfecing
            oportunidad = diff.total_seconds() / 60
            if oportunidad >= 0:
                tiempos_atencion.append(oportunidad)
        
        # Estancia (Días)
        fecha_fin = ing.ainfecegre if ing.ainfecegre else timezone.now()
        estancia = (fecha_fin - ing.ainfecing).days if ing.ainfecing else 0
        total_dias_estancia += estancia
        
        # Diagnósticos Principales
        folio_ids = list(folios.values_list('oid', flat=True))
        diags_principales = Hcndiapac.objects.using('readonly').filter(
            hcnfolio__in=folio_ids, 
            hcpdiaprin=True
        )
        
        diags_nombres = []
        for dp in diags_principales:
            d_obj = Gendiagno.objects.using('readonly').filter(oid=dp.gendiagno).first()
            if d_obj:
                nombre_d = f"{d_obj.diacodigo} - {d_obj.dianombre}"
                diags_nombres.append(nombre_d)
                conteo_diagnosticos[nombre_d] = conteo_diagnosticos.get(nombre_d, 0) + 1
        
        # Determinar si recibió atención oportuna (Ej: menos de 45 min para primer médico)
        atencion_oportuna = False
        if oportunidad is not None and oportunidad <= 45:
            atencion_oportuna = True
            
        historial_data.append({
            'ingreso': ing,
            'oportunidad': round(oportunidad, 1) if oportunidad is not None else None,
            'estancia': estancia,
            'atencion_oportuna': atencion_oportuna,
            'diagnosticos': diags_nombres,
            'is_active': ing.ainfecegre is None,
            'triage': ing.hcentriage,
            'estado_final': "Activo" if not ing.ainfecegre else "Egresado"
        })

    # Calcular promedios finales
    avg_oportunidad = sum(tiempos_atencion) / len(tiempos_atencion) if tiempos_atencion else 0
    avg_estancia = total_dias_estancia / total_ingresos if total_ingresos > 0 else 0
    
    # Top Diagnósticos
    top_diagnosticos = sorted(conteo_diagnosticos.items(), key=lambda x: x[1], reverse=True)[:5]
    
    # Calificación General
    if not tiempos_atencion:
        calificacion = "Sin Datos"
    elif avg_oportunidad <= 30:
        calificacion = "Excelente"
    elif avg_oportunidad <= 60:
        calificacion = "Aceptable"
    else:
        calificacion = "Fuera de Rango"

    context = {
        'paciente': paciente,
        'documento': documento,
        'total_ingresos': total_ingresos,
        'avg_oportunidad': round(avg_oportunidad, 1),
        'avg_estancia': round(avg_estancia, 1),
        'calificacion': calificacion,
        'historial': historial_data,
        'top_diagnosticos': top_diagnosticos,
        'searched': True,
        'found': True
    }
    
    return render(request, 'consultas/paciente_historial.html', context)
