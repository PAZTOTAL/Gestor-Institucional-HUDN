import io
import csv
from openpyxl import Workbook, load_workbook
from io import BytesIO
from django.http import HttpResponse
from django.db import transaction
from django.contrib import messages
from django.shortcuts import redirect
from django.apps import apps
from django.db import models

def get_model_safe(module_name, model_name):
    """
    Safely retrieves a model class given an app label (module_name) and model name.
    """
    try:
        return apps.get_model(module_name, model_name)
    except LookupError:
        return None

def generate_excel_template(model):
    """
    Generates an Excel template (BytesIO) for a given model.
    Includes headers for all editable fields.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = model._meta.verbose_name[:31]  # Excel limits sheet names to 31 chars

    # Determine fields to include
    headers = []
    
    # Logic to prioritize fields: 
    # 1. ForeignKeys (first, matching import logic)
    # 2. 'codigo' if exists
    # 3. 'nombre' if exists
    # 4. Other fields
    
    fk_fields = [f.name for f in model._meta.fields if f.is_relation and f.many_to_one and f.name not in ('usuario',)]
    other_fields = [f.name for f in model._meta.fields if not f.is_relation and f.name != 'id'] # ID is auto usually
    
    # Sorting for usability
    final_headers = []
    
    # Add ForeignKeys first (often dependencies)
    final_headers.extend(fk_fields)
    
    # Add rest
    final_headers.extend(other_fields)
    
    ws.append(final_headers)
    
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

def normalize_text(text):
    import re
    import unicodedata
    if not text: return ""
    # Quitar acentos
    text = "".join(c for c in unicodedata.normalize('NFD', str(text)) if unicodedata.category(c) != 'Mn')
    # Solo letras y números en minúsculas
    return re.sub(r'[^a-z0-9]', '', text.lower())

def process_excel_import(request, model, file, preview=False):
    """
    Processes an uploaded Excel or CSV file to import data into the model.
    """
    model_name_log = model._meta.model_name
    print(f"--- INICIANDO IMPORTACIÓN PARA {model_name_log} ---")
    
    rows = []
    headers = []
    
    # Check if it's a CSV
    filename = getattr(file, 'name', '').lower()
    is_csv = filename.endswith('.csv')
    
    try:
        if is_csv:
            # Leer CSV
            content = file.read().decode('latin-1')
            file.seek(0)
            import io
            import csv
            lines = content.splitlines()
            if not lines:
                return {'success': False, 'message': "Archivo vacío.", 'errors': []}
            
            delimiter = ';' if ';' in content else ','
            reader = csv.reader(io.StringIO(content), delimiter=delimiter)
            csv_rows = [[cell.strip() for cell in row] for row in list(reader)]
            
            if len(csv_rows) > 0:
                headers = csv_rows[0]
                rows = csv_rows[1:]
        else:
            # Excel
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            
    except Exception as e:
        print(f"ERROR LECTURA {model_name_log}: {e}")
        return {'success': False, 'message': f"Error al leer archivo: {e}", 'errors': []}

    print(f"ENCABEZADOS: {headers}")
    
    if not headers:
        return {'success': False, 'message': "El archivo no tiene encabezados.", 'errors': []}

    # Map headers to fields
    valid_map = {} # col_idx -> field_name
    
    # Pre-normalizar campos del modelo
    model_fields = {f.name: f for f in model._meta.get_fields() if hasattr(f, 'name')}
    norm_to_real = {normalize_text(f.name): f.name for f in model_fields.values()}
    verbose_to_real = {normalize_text(f.verbose_name): f.name for f in model_fields.values() if hasattr(f, 'verbose_name') and f.verbose_name}
    
    for idx, h in enumerate(headers):
        if not h: continue
        h_norm = normalize_text(h)
        if not h_norm: continue
        
        if h_norm in norm_to_real:
            valid_map[idx] = norm_to_real[h_norm]
        elif h_norm in verbose_to_real:
            valid_map[idx] = verbose_to_real[h_norm]
        else:
            # Variaciones comunes
            variations = {
                'codigo': 'codigo_formato',
                'nombre': 'nombre_formato',
                'elaborado': 'elaborado_por',
                'modificado': 'modificado_por',
                'version': 'version'
            }
            if h_norm in variations:
                valid_map[idx] = variations[h_norm]
            
    print(f"MAPA FINAL {model_name_log}: {valid_map}")
    
    if not valid_map:
        return {'success': False, 'message': f"No se encontraron columnas válidas. Esperadas: {list(verbose_to_real.keys())[:5]}...", 'errors': []}

    errors = []
    created_count = 0
    
    try:
        for row_idx, row in enumerate(rows, start=2):
            if not any(row): continue
            
            row_data = {}
            row_errors = []
            
            for col_idx, field_name in valid_map.items():
                val = row[col_idx]
                field = model_fields.get(field_name)
                if not field: continue
                
                if isinstance(val, str):
                    val = val.strip()
                if val == "" or val is None:
                    continue # Skip empty fields, let defaults handle it
                    
                # FK Resolving logic
                if field.is_relation and field.many_to_one:
                    related = field.related_model
                    fk_obj = None
                    
                    # Try 'codigo'
                    if hasattr(related, 'codigo'):
                        try:
                            fk_obj = related.objects.get(codigo=val)
                        except: pass
                    
                    # Try 'username' (common for User models)
                    if not fk_obj and hasattr(related, 'username'):
                        try:
                            fk_obj = related.objects.get(username=val)
                        except: pass
                        
                    # Try 'nombre'
                    if not fk_obj and hasattr(related, 'nombre'):
                        try:
                            fk_obj = related.objects.get(nombre=val)
                        except: pass
                    
                    # Try PK
                    if not fk_obj:
                        try:
                            fk_obj = related.objects.get(pk=val)
                        except: pass
                            
                    if fk_obj:
                        val = fk_obj
                    else:
                        row_errors.append(f"No se halló '{field.verbose_name}' con valor '{val}'")
                        continue
                
                # Truncate strings if they exceed CharField max_length
                if val and isinstance(val, str) and isinstance(field, models.CharField):
                    if field.max_length and len(val) > field.max_length:
                        val = val[:field.max_length]

                row_data[field_name] = val
            
            if row_errors:
                errors.append({'fila': row_idx, 'error': "; ".join(row_errors)})
                continue

            try:
                with transaction.atomic():
                    instance = model(**row_data)
                    
                    # Auto-fill user if model has 'usuario' or 'elaborado_por' (if empty)
                    if hasattr(instance, 'usuario') and not instance.usuario_id:
                         instance.usuario = request.user
                    
                    # Specific to Formatos_Hudn / Organigrama where elaborated_por is common
                    if hasattr(instance, 'elaborado_por_id') and not instance.elaborado_por_id:
                        instance.elaborado_por = request.user
                    
                    instance.full_clean()
                    
                    if preview:
                        # Emular guardado para validar BD pero revertir
                        transaction.set_rollback(True)
                    else:
                        instance.save()
                        
                    created_count += 1
            except Exception as e:
                errors.append({'fila': row_idx, 'error': str(e)})

    except Exception as e:
        return {'success': False, 'message': f"Error crítico: {e}", 'errors': []}

    # Generate Error CSV if needed
    if errors:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="errores_{model._meta.model_name}.csv"'
        writer = csv.writer(response)
        writer.writerow(['Fila', 'Error'])
        for err in errors:
            writer.writerow([err['fila'], err['error']])
        
        return {
            'success': False, 
            'message': f"Se encontraron {len(errors)} errores.", 
            'errors': errors, 
            'response': response,
            'preview_count': created_count if preview else 0
        }

    return {
        'success': True, 
        'message': f"Se procesaron {created_count} registros.", 
        'count': created_count
    }
